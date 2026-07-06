import io
import unittest
import urllib.error
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

from src import comptox_use


COMPT0X_RANKING_COLUMNS = (
    "产品场景Top5",
    "功能用途Top5",
    "综合候选Top5",
    "前五用途",
    "用途来源",
    "用途1",
    "用途1_英文证据",
    "用途1_证据数量",
    "产品场景用途1",
    "产品场景用途1_英文证据",
    "产品场景用途1_证据数量",
    "功能用途1",
    "功能用途1_英文证据",
    "功能用途1_证据数量",
)


def _candidate(source_type, raw_use="cleaning agent", use_cn="清洁用品", **extra):
    return {
        "source_type": source_type,
        "source": f"dashboard:{source_type}",
        "raw_use": raw_use,
        "use_cn": use_cn,
        "general_category": "cleaning",
        "product_family": "",
        "product_type": "",
        "reported_use": "",
        "harmonized_use": "",
        "evidence_count": 1,
        "description": "",
        "specificity": 1,
        **extra,
    }


class CompToxDashboardModeTests(unittest.TestCase):
    def test_dashboard_mode_skips_unconfigured_api(self):
        with (
            patch.object(
                comptox_use,
                "_api_get_json",
                side_effect=AssertionError("the disabled API must not be called"),
            ),
            patch.object(comptox_use, "_dashboard_get_html", return_value="page"),
            patch.object(
                comptox_use,
                "_extract_dashboard_product_categories",
                return_value=[_candidate("product_category")],
            ),
            patch.object(
                comptox_use,
                "_extract_dashboard_functional_uses",
                return_value=[_candidate("functional_use")],
            ),
        ):
            candidates, warnings = comptox_use.fetch_use_candidates(
                "DTXSID0020153", api_base="", dashboard_fallback=True
            )

        self.assertEqual(warnings, [])
        self.assertEqual(
            {candidate["source_type"] for candidate in candidates},
            {"product_category", "functional_use"},
        )

    def test_dashboard_mode_resolves_dtxsid_without_api(self):
        record = {
            "dtxsid": "DTXSID0020153",
            "preferredName": "Benzyl chloride",
            "casrn": "100-44-7",
        }
        with (
            patch.object(
                comptox_use,
                "_api_get_json",
                side_effect=AssertionError("the disabled API must not be called"),
            ),
            patch.object(
                comptox_use,
                "_dashboard_search_chemical_candidates",
                return_value=[record],
            ),
        ):
            result = comptox_use.resolve_dtxsid(
                pd.Series({"cas": "100-44-7", "compound": "Benzyl chloride"}),
                api_base="",
            )

        self.assertEqual(result["dtxsid"], "DTXSID0020153")
        self.assertEqual(result["status"], "通过 Dashboard cas 匹配")

    def test_batch_surfaces_scope_note_instead_of_api_failures(self):
        with (
            patch.object(
                comptox_use,
                "fetch_use_candidates",
                return_value=([_candidate("product_category")], []),
            ),
        ):
            summary_df, _, errors_df = comptox_use.run_comptox_use_batch(
                pd.DataFrame(
                    [{"compound": "Benzyl chloride", "dtxsid": "DTXSID0020153"}]
                ),
                api_base="",
                delay_seconds=0,
            )

        self.assertTrue(errors_df.empty)
        self.assertEqual(
            summary_df.loc[0, "query_notes"], comptox_use.DASHBOARD_ONLY_QUERY_NOTE
        )

    def test_batch_summary_separates_product_and_functional_uses(self):
        candidates = [
            _candidate(
                "product_category",
                raw_use="Personal care products",
                use_cn="个人护理用品",
            ),
            _candidate(
                "functional_use",
                raw_use="Fragrance",
                use_cn="香精香料",
            ),
        ]
        with (
            patch.object(comptox_use, "fetch_use_candidates", return_value=(candidates, [])),
        ):
            summary_df, _, _ = comptox_use.run_comptox_use_batch(
                pd.DataFrame(
                    [{"compound": "Benzyl chloride", "dtxsid": "DTXSID0020153"}]
                ),
                api_base="",
                delay_seconds=0,
            )

        self.assertEqual(summary_df.loc[0, "产品用途类别"], "Personal care products (1)")
        self.assertEqual(summary_df.loc[0, "已收集化学功能用途"], "香精香料 (Fragrance)")
        self.assertTrue(pd.isna(summary_df.loc[0, "预测化学功能用途"]))
        self.assertIn("product-use-categories/DTXSID0020153", summary_df.loc[0, "CompTox来源链接"])
        self.assertIn("chemical-functional-use/DTXSID0020153", summary_df.loc[0, "CompTox来源链接"])
        self.assertNotIn("CompTox产品用途页面", summary_df.columns)
        self.assertNotIn("CompTox功能用途页面", summary_df.columns)

    def test_batch_summary_omits_top_ranking_columns(self):
        candidates = [
            _candidate(
                "product_category",
                raw_use="Personal care products",
                use_cn="个人护理用品",
                evidence_count=12,
            ),
            _candidate(
                "product_category",
                raw_use="Cleaning products",
                use_cn="清洁用品",
                evidence_count=3,
            ),
            _candidate(
                "functional_use",
                raw_use="Fragrance",
                use_cn="芳香剂",
                probability=0.91,
                functional_use_source="predicted",
                evidence_count=0.91,
            ),
        ]
        with patch.object(comptox_use, "fetch_use_candidates", return_value=(candidates, [])):
            summary_df, _, _ = comptox_use.run_comptox_use_batch(
                pd.DataFrame([{"compound": "Benzyl chloride", "dtxsid": "DTXSID0020153"}]),
                api_base="",
                delay_seconds=0,
            )

        self.assertEqual(summary_df.loc[0, "产品用途类别"], "Personal care products (12)；Cleaning products (3)")
        self.assertEqual(summary_df.loc[0, "预测化学功能用途"], "芳香剂 (Fragrance, p=0.910)")
        for column in COMPT0X_RANKING_COLUMNS:
            self.assertNotIn(column, summary_df.columns)

    def test_product_summary_keeps_all_distinct_english_puc_scenarios(self):
        candidates = [
            _candidate(
                "product_category",
                raw_use="Cleaning products and household care:air freshener",
                use_cn="清洁用品",
                evidence_count=32,
            ),
            _candidate(
                "product_category",
                raw_use="Cleaning products and household care:shoes:shoe polish or protectant",
                use_cn="清洁用品",
                evidence_count=2,
            ),
        ]
        with patch.object(comptox_use, "fetch_use_candidates", return_value=(candidates, [])):
            summary_df, _, _ = comptox_use.run_comptox_use_batch(
                pd.DataFrame([{"compound": "p-Cymene", "dtxsid": "DTXSID3026645"}]),
                api_base="",
                delay_seconds=0,
            )

        self.assertEqual(summary_df.loc[0, "产品用途类别"], "Cleaning products and household care:air freshener (32)；Cleaning products and household care:shoes:shoe polish or protectant (2)")
        self.assertNotIn("产品场景用途1", summary_df.columns)

    def test_functional_summary_uses_predicted_only_and_preserves_reported_detail(self):
        candidates = [
            _candidate(
                "functional_use",
                raw_use="Flavouring and nutrient",
                use_cn="调味剂",
                evidence_count=3,
                functional_use_source="reported",
                probability=pd.NA,
            ),
            _candidate(
                "functional_use",
                raw_use="flavorant",
                use_cn="调味剂",
                evidence_count=0.2731,
                functional_use_source="predicted",
                probability=0.2731,
            ),
        ]
        with patch.object(comptox_use, "fetch_use_candidates", return_value=(candidates, [])):
            summary_df, candidates_df, _ = comptox_use.run_comptox_use_batch(
                pd.DataFrame([{"compound": "p-Cymene", "dtxsid": "DTXSID3026645"}]),
                api_base="",
                delay_seconds=0,
            )

        self.assertEqual(summary_df.loc[0, "已收集化学功能用途"], "调味剂 (Flavouring and nutrient)")
        self.assertEqual(summary_df.loc[0, "预测化学功能用途"], "调味剂 (flavorant, p=0.273)")
        for column in COMPT0X_RANKING_COLUMNS:
            self.assertNotIn(column, summary_df.columns)

        functional_df = comptox_use.build_functional_use_table(candidates_df)
        self.assertCountEqual(list(functional_df["来源类型"]), ["reported", "predicted"])
        self.assertCountEqual(
            list(zip(functional_df["来源类型"], functional_df["英文功能用途"])),
            [("reported", "Flavouring and nutrient"), ("predicted", "flavorant")],
        )

    def test_product_use_table_preserves_puc_hierarchy(self):
        candidates_df = pd.DataFrame(
            [
                {
                    "compound": "Example",
                    "dtxsid": "DTXSID0000001",
                    "source_type": "product_category",
                    "source": "dashboard:product_category",
                    "raw_use": "Personal Care Products",
                    "use_cn": "个人护理用品",
                    "general_category": "Consumer Products",
                    "product_family": "Personal Care",
                    "product_type": "Fragrance",
                    "evidence_count": 7,
                    "description": "Products used for personal care.",
                }
            ]
        )

        product_df = comptox_use.build_product_use_table(candidates_df)

        self.assertEqual(product_df.loc[0, "产品用途类别"], "Personal Care Products")
        self.assertEqual(product_df.loc[0, "英文产品用途类别"], "Personal Care Products")
        self.assertEqual(product_df.loc[0, "general_category"], "Consumer Products")
        self.assertEqual(product_df.loc[0, "product_family"], "Personal Care")
        self.assertEqual(product_df.loc[0, "product_type"], "Fragrance")
        self.assertEqual(product_df.loc[0, "product_count"], 7)
        self.assertIn("product-use-categories/DTXSID0000001", product_df.loc[0, "CompTox产品用途链接"])

    def test_dashboard_functional_use_extracts_predicted_probability_table(self):
        html = """
        <script>
        window.__NUXT__=(function(){
          return {data:[{cfuData:{
            reportedFunctionalUse:[],
            predictedFunctionalUse:[
              {harmonizedFunctionalUse:"fragrance",probability:.9126},
              {harmonizedFunctionalUse:"antioxidant",probability:.3679}
            ]
          }}]}
        })()
        </script>
        """

        candidates = comptox_use._extract_dashboard_functional_uses(html)

        self.assertEqual([item["raw_use"] for item in candidates], ["fragrance", "antioxidant"])
        self.assertEqual(candidates[0]["functional_use_source"], "predicted")
        self.assertAlmostEqual(candidates[0]["probability"], 0.9126)

    def test_reported_functional_use_synonyms_count_as_single_evidence_item(self):
        candidates = comptox_use._extract_functional_use_candidates(
            [
                {
                    "harmonizedFunctionalUse": "Intermediate",
                    "reportedFunctionalUse": "intermediate",
                },
                {
                    "harmonizedFunctionalUse": "Intermediate",
                    "reportedFunctionalUse": "intermediates",
                },
            ],
            "dashboard:functional_use",
        )

        self.assertEqual(len(candidates), 1)
        self.assertEqual(candidates[0]["functional_use_source"], "reported")
        self.assertEqual(candidates[0]["reported_use"], "intermediate | intermediates")
        self.assertEqual(candidates[0]["evidence_count"], 1)

    def test_functional_use_extraction_separates_reported_and_predicted_for_same_label(self):
        candidates = comptox_use._extract_functional_use_candidates(
            [
                {
                    "harmonizedFunctionalUse": "Intermediate",
                    "reportedFunctionalUse": "intermediate",
                },
                {
                    "harmonizedFunctionalUse": "Intermediate",
                    "probability": 0.72,
                },
            ],
            "dashboard:functional_use",
        )

        by_source = {candidate["functional_use_source"]: candidate for candidate in candidates}
        self.assertEqual(set(by_source), {"reported", "predicted"})
        self.assertEqual(by_source["reported"]["evidence_count"], 1)
        self.assertTrue(pd.isna(by_source["reported"]["probability"]))
        self.assertAlmostEqual(by_source["predicted"]["evidence_count"], 0.72)
        self.assertAlmostEqual(by_source["predicted"]["probability"], 0.72)

    def test_functional_use_extraction_splits_combined_reported_and_predicted_record(self):
        candidates = comptox_use._extract_functional_use_candidates(
            [
                {
                    "harmonizedFunctionalUse": "fragrance",
                    "reportedFunctionalUse": "Fragrance",
                    "probability": 0.91,
                }
            ],
            "dashboard:functional_use",
        )

        by_source = {candidate["functional_use_source"]: candidate for candidate in candidates}
        self.assertEqual(set(by_source), {"reported", "predicted"})
        self.assertEqual(by_source["reported"]["reported_use"], "Fragrance")
        self.assertEqual(by_source["reported"]["evidence_count"], 1)
        self.assertTrue(pd.isna(by_source["reported"]["probability"]))
        self.assertEqual(by_source["predicted"]["reported_use"], "")
        self.assertAlmostEqual(by_source["predicted"]["evidence_count"], 0.91)
        self.assertAlmostEqual(by_source["predicted"]["probability"], 0.91)

    def test_functional_use_translation_handles_predicted_use_labels(self):
        self.assertEqual(comptox_use.classify_use_cn("fragrance"), "芳香剂")
        self.assertEqual(comptox_use.classify_use_cn("flavorant"), "调味剂")
        self.assertEqual(comptox_use.classify_use_cn("flame_retardant"), "阻燃剂")
        self.assertEqual(comptox_use.classify_use_cn("antimicrobial"), "抗微生物剂")
        self.assertEqual(comptox_use.classify_use_cn("skin_protectant"), "皮肤保护剂")
        self.assertEqual(comptox_use.classify_use_cn("skin_conditioner"), "皮肤调理剂")
        self.assertEqual(comptox_use.classify_use_cn("specialty_unmapped_use"), "")

    def test_functional_use_table_displays_unmapped_chinese_label_as_other_use(self):
        candidates_df = pd.DataFrame(
            [
                {
                    "compound": "Example",
                    "dtxsid": "DTXSID0000001",
                    "source_type": "functional_use",
                    "source": "dashboard:functional_use",
                    "raw_use": "specialty_unmapped_use",
                    "use_cn": comptox_use.classify_use_cn("specialty_unmapped_use"),
                    "reported_use": "",
                    "harmonized_use": "specialty_unmapped_use",
                    "evidence_count": 0.42,
                    "probability": 0.42,
                    "functional_use_source": "predicted",
                }
            ]
        )

        functional_df = comptox_use.build_functional_use_table(candidates_df)

        self.assertEqual(functional_df.loc[0, "功能用途"], "其他用途：specialty_unmapped_use")
        self.assertEqual(functional_df.loc[0, "英文功能用途"], "specialty_unmapped_use")

    def test_functional_use_table_can_filter_predicted_and_reported_rows(self):
        candidates_df = pd.DataFrame(
            [
                {
                    "compound": "Example",
                    "dtxsid": "DTXSID0000001",
                    "source_type": "functional_use",
                    "source": "dashboard:functional_use",
                    "raw_use": "fragrance",
                    "use_cn": "芳香剂",
                    "reported_use": "",
                    "harmonized_use": "fragrance",
                    "evidence_count": 0.91,
                    "probability": 0.91,
                    "functional_use_source": "predicted",
                },
                {
                    "compound": "Example",
                    "dtxsid": "DTXSID0000001",
                    "source_type": "functional_use",
                    "source": "dashboard:functional_use",
                    "raw_use": "Fragrance",
                    "use_cn": "芳香剂",
                    "reported_use": "Fragrance",
                    "harmonized_use": "Fragrance",
                    "evidence_count": 1,
                    "probability": pd.NA,
                    "functional_use_source": "reported",
                },
            ]
        )

        predicted_df = comptox_use.build_functional_use_table(candidates_df, functional_source="predicted")
        reported_df = comptox_use.build_functional_use_table(candidates_df, functional_source="reported")

        self.assertEqual(predicted_df["来源类型"].tolist(), ["predicted"])
        self.assertEqual(reported_df["来源类型"].tolist(), ["reported"])
        self.assertEqual(predicted_df.loc[0, "reported_use"], "")
        self.assertEqual(reported_df.loc[0, "reported_use"], "Fragrance")

    def test_batch_summary_shows_predicted_functional_use_probability(self):
        input_df = pd.DataFrame([{"compound": "2CB", "dtxsid": "DTXSID2069284"}])
        candidates = [
            _candidate(
                "functional_use",
                raw_use="fragrance",
                use_cn="芳香剂",
                probability=0.9126,
                functional_use_source="predicted",
                evidence_count=0.9126,
            ),
            _candidate(
                "functional_use",
                raw_use="antioxidant",
                use_cn="抗氧化剂",
                probability=0.3679,
                functional_use_source="predicted",
                evidence_count=0.3679,
            ),
        ]
        with (
            patch.object(comptox_use, "fetch_use_candidates", return_value=(candidates, [])),
        ):
            summary_df, candidates_df, _ = comptox_use.run_comptox_use_batch(
                input_df,
                api_base="",
                delay_seconds=0,
            )

        self.assertTrue(pd.isna(summary_df.loc[0, "已收集化学功能用途"]))
        self.assertEqual(
            summary_df.loc[0, "预测化学功能用途"],
            "芳香剂 (fragrance, p=0.913)；抗氧化剂 (antioxidant, p=0.368)",
        )
        self.assertIn("probability", candidates_df.columns)
        self.assertAlmostEqual(candidates_df.loc[0, "probability"], 0.9126)

        functional_df = comptox_use.build_functional_use_table(candidates_df)
        self.assertEqual(
            list(functional_df["英文功能用途"]),
            ["fragrance", "antioxidant"],
        )
        self.assertEqual(list(functional_df["来源类型"]), ["predicted", "predicted"])
        self.assertAlmostEqual(functional_df.loc[0, "预测概率"], 0.9126)

        workbook = comptox_use.build_result_workbook(
            input_df,
            summary_df=summary_df,
            candidates_df=candidates_df,
            errors_df=pd.DataFrame(),
        )
        book = load_workbook(io.BytesIO(workbook.getvalue()), read_only=True)
        self.assertIn("Use_Summary", book.sheetnames)
        self.assertNotIn("Top5_Use_Summary", book.sheetnames)
        self.assertNotIn("Functional_Uses", book.sheetnames)
        self.assertIn("Functional_Uses_Predicted", book.sheetnames)
        self.assertIn("Functional_Uses_Reported", book.sheetnames)
        self.assertIn("Product_Use_Categories", book.sheetnames)
        self.assertIn("Evidence_Metadata", book.sheetnames)
        self.assertEqual(
            [cell.value for cell in next(book["Warnings"].iter_rows(min_row=1, max_row=1))],
            comptox_use.WARNING_COLUMNS,
        )
        metadata_rows = list(book["Evidence_Metadata"].iter_rows(values_only=True))
        self.assertIn("product_category", [row[0] for row in metadata_rows[1:]])
        self.assertIn("functional_use", [row[0] for row in metadata_rows[1:]])

    def test_dashboard_request_retries_transient_network_failure(self):
        class Response:
            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc_value, traceback):
                return False

            def read(self):
                return b"dashboard page"

        calls = []

        def urlopen(*args, **kwargs):
            calls.append((args, kwargs))
            if len(calls) == 1:
                raise urllib.error.URLError("connection reset")
            return Response()

        with (
            patch.object(comptox_use.urllib.request, "urlopen", side_effect=urlopen),
            patch.object(comptox_use.time, "sleep") as sleep,
        ):
            page = comptox_use._dashboard_get_html("chemical/example", timeout=1)

        self.assertEqual(page, "dashboard page")
        self.assertEqual(len(calls), 2)
        sleep.assert_called_once_with(comptox_use.DASHBOARD_RETRY_DELAY_SECONDS)


if __name__ == "__main__":
    unittest.main()
