import copy
import json
import math
import tempfile
import unittest
import urllib.parse
import zipfile
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

from src import episuite_io
from src.batch_runner import BatchResult
from src.query_cache import cache_control, use_cache_path


ETHANOL_CAS_AND_SMILES_RESPONSE = {
    "chemicalProperties": {
        "cas": "000064-17-5",
        "smiles": "CCO",
        "name": "ETHANOL",
        "molecularFormula": "C2 H6 O1",
        "molecularWeight": 46.07,
    },
    "parameters": {
        "cas": "64-17-5",
        "smiles": "CCO",
    },
    "logKow": {
        "selectedValue": {"value": -0.31, "units": None, "valueType": "EXPERIMENTAL"},
        "estimatedValue": {"value": -0.1411999762058258, "units": "", "valueType": "ESTIMATED"},
        "experimentalValue": None,
    },
    "waterSolubilityFromWaterNt": {
        "selectedValue": {"value": 1000000.0, "units": "mg/L", "valueType": "EXPERIMENTAL"},
        "estimatedValue": {"value": 452462.28125, "units": "mg/L", "valueType": "ESTIMATED"},
        "experimentalValue": None,
    },
    "waterSolubilityFromLogKow": {
        "selectedValue": {"value": 1000000.0, "units": "mg/L", "valueType": "EXPERIMENTAL"},
        "estimatedValue": {"value": 857740.375, "units": "mg/L", "valueType": "ESTIMATED"},
        "experimentalValue": None,
    },
    "vaporPressure": {
        "selectedValue": {"value": 59.3, "units": "mmHg", "valueType": "EXPERIMENTAL"},
        "estimatedValue": {"value": 70.6, "units": "mmHg", "valueType": "ESTIMATED"},
        "experimentalValue": None,
    },
    "henrysLawConstant": {
        "selectedValue": {"value": 5.0e-6, "units": "atm-m3/mol", "valueType": "EXPERIMENTAL"},
        "estimatedValue": {"value": 6.1e-6, "units": "atm-m3/mol", "valueType": "ESTIMATED"},
        "experimentalValue": None,
    },
    "logKoc": {
        "selectedValue": {"value": 1.2, "units": "L/kg", "valueType": "EXPERIMENTAL"},
        "estimatedValue": {"value": 1.05, "units": "L/kg", "valueType": "ESTIMATED"},
        "experimentalValue": None,
    },
    "biodegradationRate": {
        "models": [
            {"name": "Ultimate Biodegradation Timeframe", "value": 2.1, "description": "weeks"},
            {"name": "MITI Linear Model Prediction", "value": 0.72, "description": "readily biodegradable"},
        ],
    },
    "atmosphericHalfLife": {
        "estimatedValue": {"value": 10.2, "units": "hours", "valueType": "ESTIMATED"},
        "estimatedHydroxylRadicalReactionRateConstant": {
            "value": 3.3e-12,
            "units": "cm3/molecule-sec",
            "valueType": "ESTIMATED",
        },
    },
    "bioconcentration": {
        "bioconcentrationFactor": 3.2,
        "logBioconcentrationFactor": 0.51,
        "bioaccumulationFactor": 4.8,
        "logBioaccumulationFactor": 0.68,
    },
    "sewageTreatmentModel": {
        "model": {
            "TotalRemoval": {"Percent": 88.1},
            "FinalEffluent": {"Percent": 11.9},
        },
    },
    "fugacityModel": {
        "model": {
            "Air": [{"MassAmount": 12.0, "HalfLife": 8.0}],
            "Water": [{"MassAmount": 70.0, "HalfLife": 180.0}],
            "Soil": [{"MassAmount": 15.0, "HalfLife": 300.0}],
            "Sediment": [{"MassAmount": 3.0, "HalfLife": 900.0}],
            "Persistence": 240.0,
        },
    },
    "waterVolatilization": {
        "riverHalfLifeHours": 4.5,
        "lakeHalfLifeHours": 33.0,
    },
    "ecosar": {
        "modelResults": [
            {
                "className": "Neutral Organics",
                "organism": "Fish",
                "duration": "96-hr",
                "endpoint": "LC50",
                "concentration": 13000.0,
                "units": "mg/L",
                "maxLogKow": 5.0,
                "warnings": ["Above solubility limit"],
            },
            {
                "className": "Neutral Organics",
                "organism": "Daphnid",
                "duration": "48-hr",
                "endpoint": "LC50",
                "concentration": 12000.0,
                "units": "mg/L",
            },
        ],
    },
}


class EPISuiteCasValueTests(unittest.TestCase):
    def _response_with_koawin_model(self):
        response = copy.deepcopy(ETHANOL_CAS_AND_SMILES_RESPONSE)
        kow = 10.0 ** response["logKow"]["estimatedValue"]["value"]
        kaw = 0.001
        koa = kow / kaw
        response["logKoa"] = {
            "selectedValue": {
                "value": math.log10(koa),
                "units": "",
                "valueType": "ESTIMATED",
            },
            "estimatedValue": {
                "value": math.log10(koa),
                "units": "",
                "valueType": "ESTIMATED",
                "model": {
                    "kow": kow,
                    "kaw": kaw,
                    "koa": koa,
                    "logKoa": math.log10(koa),
                },
            },
            "experimentalValues": [],
        }
        return response

    def test_name_alias_without_smiles_is_valid_epi_input(self):
        normalized = episuite_io.normalize_input_columns(
            pd.DataFrame({"name": [" Ethanol "]})
        )

        valid, message = episuite_io.validate_input(normalized)

        self.assertTrue(valid, message)
        self.assertEqual(normalized.loc[0, "compound"], "Ethanol")
        self.assertNotIn("smiles", normalized.columns)

    def test_name_only_input_zip_contains_query_terms_without_blank_smiles(self):
        package = episuite_io.build_input_zip(
            pd.DataFrame({"compound": ["Ethanol", "Benzene"], "smiles": [pd.NA, "c1ccccc1"]})
        )

        with zipfile.ZipFile(package) as archive:
            self.assertEqual(
                archive.read("episuite_query_terms.txt").decode("utf-8"),
                "Ethanol\nc1ccccc1\n",
            )
            self.assertEqual(
                archive.read("episuite_smiles_only.txt").decode("utf-8"),
                "c1ccccc1\n",
            )
            self.assertIn(
                "Ethanol",
                archive.read("episuite_input.csv").decode("utf-8"),
            )

    @patch("src.episuite_io.run_ordered_batch", return_value=[])
    def test_epi_batch_enables_three_transient_failure_rounds(self, runner):
        episuite_io.run_epi_web_batch(
            pd.DataFrame({"compound": ["A"], "smiles": ["CCO"]}),
            delay_seconds=0,
        )

        options = runner.call_args.kwargs
        self.assertEqual(options["max_attempts"], 3)
        should_retry = options["should_retry"]
        self.assertTrue(
            should_retry(BatchResult(index=0, error=RuntimeError("HTTP 429: busy")))
        )
        self.assertTrue(
            should_retry(
                BatchResult(
                    index=0,
                    value=(
                        pd.DataFrame(),
                        pd.DataFrame(),
                        pd.DataFrame({"error": ["connection reset by peer"]}),
                    ),
                )
            )
        )
        self.assertFalse(
            should_retry(BatchResult(index=0, error=RuntimeError("HTTP 400: bad input")))
        )

    def test_normalize_input_columns_keeps_optional_cas(self):
        df = episuite_io.normalize_input_columns(
            pd.DataFrame(
                {
                    "Compound": ["Ethanol"],
                    "Canonical_SMILES": [" CCO "],
                    "CASRN": [" 64-17-5 "],
                }
            )
        )

        self.assertEqual(df.loc[0, "compound"], "Ethanol")
        self.assertEqual(df.loc[0, "smiles"], "CCO")
        self.assertEqual(df.loc[0, "cas"], "64-17-5")

    def test_normalize_input_treats_literal_null_smiles_as_missing(self):
        normalized = episuite_io.normalize_input_columns(
            pd.DataFrame({"compound": ["Bad"], "smiles": ["null"]})
        )

        self.assertTrue(pd.isna(normalized.loc[0, "smiles"]))

    @patch("src.episuite_io.urllib.request.urlopen")
    def test_call_epi_web_api_sends_cas_and_smiles_when_cas_is_present(self, urlopen):
        response = unittest.mock.MagicMock()
        response.read.return_value = json.dumps({"ok": True}).encode("utf-8")
        urlopen.return_value.__enter__.return_value = response

        with cache_control(False):
            episuite_io.call_epi_web_api("CCO", cas="64-17-5", api_url="https://example.test/api/submit")

        request = urlopen.call_args.args[0]
        self.assertIn("smiles=CCO", request.full_url)
        self.assertIn("cas=64-17-5", request.full_url)

    @patch("src.episuite_io.urllib.request.urlopen")
    def test_call_epi_web_api_reuses_cached_response(self, urlopen):
        response = unittest.mock.MagicMock()
        response.read.return_value = json.dumps({"ok": True}).encode("utf-8")
        urlopen.return_value.__enter__.return_value = response

        with tempfile.TemporaryDirectory() as tmpdir:
            with use_cache_path(Path(tmpdir) / "queries.sqlite3"):
                first = episuite_io.call_epi_web_api(
                    "CCO",
                    cas="64-17-5",
                    api_url="https://example.test/api/submit",
                )
                second = episuite_io.call_epi_web_api(
                    "CCO",
                    cas="64-17-5",
                    api_url="https://example.test/api/submit",
                )

        self.assertEqual(first, {"ok": True})
        self.assertEqual(second, {"ok": True})
        urlopen.assert_called_once()

    @patch("src.episuite_io.urllib.request.urlopen")
    def test_call_epi_web_search_uses_sibling_search_endpoint(self, urlopen):
        response = unittest.mock.MagicMock()
        response.read.return_value = json.dumps(
            [{"name": "ETHANOL", "smiles": "OCC", "cas": "000064-17-5"}]
        ).encode("utf-8")
        urlopen.return_value.__enter__.return_value = response

        with cache_control(False):
            candidates = episuite_io.call_epi_web_search(
                " Ethanol ",
                api_url="https://example.test/api/submit",
            )

        request = urlopen.call_args.args[0]
        self.assertIn("/api/search?", request.full_url)
        self.assertIn("query=Ethanol", request.full_url)
        self.assertEqual(candidates[0]["cas"], "000064-17-5")

    @patch("src.episuite_io.urllib.request.urlopen")
    def test_call_epi_web_search_merges_params_before_existing_fragment(self, urlopen):
        response = unittest.mock.MagicMock()
        response.read.return_value = b"[]"
        urlopen.return_value.__enter__.return_value = response

        with cache_control(False):
            episuite_io.call_epi_web_search(
                "Ethanol",
                api_url="https://example.test/api/submit?existing=base#kept",
                limit=42,
            )

        parsed = urllib.parse.urlsplit(urlopen.call_args.args[0].full_url)
        self.assertEqual(parsed.path, "/api/search")
        self.assertEqual(
            urllib.parse.parse_qs(parsed.query),
            {"existing": ["base"], "query": ["Ethanol"], "limit": ["42"]},
        )
        self.assertEqual(parsed.fragment, "kept")

    @patch("src.episuite_io.call_epi_web_search")
    def test_exact_name_resolution_ignores_case_and_uses_first_exact_candidate(self, search):
        search.return_value = [
            {"name": "Ethanol derivative", "smiles": "CCC", "cas": "1-11-1"},
            {"name": " ETHANOL ", "smiles": "OCC", "cas": "000064-17-5"},
            {"name": "ethanol", "smiles": "CCO", "cas": "64-17-5"},
        ]

        resolved = episuite_io.resolve_epi_name_exact("Ethanol")

        self.assertEqual(resolved["name"], "ETHANOL")
        self.assertEqual(resolved["smiles"], "OCC")
        self.assertEqual(resolved["cas"], "000064-17-5")

    @patch("src.episuite_io.call_epi_web_search")
    def test_exact_name_resolution_rejects_fuzzy_only_candidates(self, search):
        search.return_value = [
            {"name": "Ethanol derivative", "smiles": "CCC", "cas": "1-11-1"}
        ]

        with self.assertRaisesRegex(RuntimeError, "完全一致"):
            episuite_io.resolve_epi_name_exact("Ethanol")

    @patch("src.episuite_io.call_epi_web_search")
    def test_exact_name_resolution_skips_malformed_candidates(self, search):
        search.return_value = ["malformed", {"name": "Ethanol derivative", "smiles": "CCC"}]

        with self.assertRaisesRegex(RuntimeError, "完全一致"):
            episuite_io.resolve_epi_name_exact("Ethanol")

    @patch("src.episuite_io.call_epi_web_search")
    def test_exact_name_resolution_requires_candidate_smiles(self, search):
        search.return_value = [{"name": "ETHANOL", "smiles": "", "cas": "64-17-5"}]

        with self.assertRaisesRegex(RuntimeError, "SMILES"):
            episuite_io.resolve_epi_name_exact("Ethanol")

    @patch("src.episuite_io.urllib.request.urlopen")
    def test_call_epi_web_search_reuses_cached_response(self, urlopen):
        response = unittest.mock.MagicMock()
        response.read.return_value = b"[]"
        urlopen.return_value.__enter__.return_value = response

        with tempfile.TemporaryDirectory() as tmpdir:
            with use_cache_path(Path(tmpdir) / "queries.sqlite3"):
                first = episuite_io.call_epi_web_search("Ethanol")
                second = episuite_io.call_epi_web_search("Ethanol")

        self.assertEqual(first, [])
        self.assertEqual(second, [])
        urlopen.assert_called_once()

    def test_extract_epi_web_summary_keeps_selected_estimated_and_experimental_values(self):
        summary = episuite_io.extract_epi_web_summary(
            "Ethanol",
            "CCO",
            ETHANOL_CAS_AND_SMILES_RESPONSE,
            cas="64-17-5",
        )

        self.assertEqual(summary["cas"], "64-17-5")
        self.assertEqual(summary["epi_cas"], "000064-17-5")
        self.assertEqual(summary["log_kow"], -0.31)
        self.assertEqual(summary["log_kow_selected"], -0.31)
        self.assertEqual(summary["log_kow_estimated"], -0.1411999762058258)
        self.assertEqual(summary["log_kow_experimental"], -0.31)
        self.assertEqual(summary["water_solubility_mg_l"], 1000000.0)
        self.assertEqual(summary["water_solubility_selected"], 1000000.0)
        self.assertEqual(summary["water_solubility_estimated"], 452462.28125)
        self.assertEqual(summary["water_solubility_experimental"], 1000000.0)
        self.assertEqual(summary["vapor_pressure_selected"], 59.3)
        self.assertEqual(summary["vapor_pressure_estimated"], 70.6)
        self.assertEqual(summary["vapor_pressure_experimental"], 59.3)
        self.assertEqual(summary["henry_selected"], 5.0e-6)
        self.assertEqual(summary["henry_estimated"], 6.1e-6)
        self.assertEqual(summary["henry_experimental"], 5.0e-6)
        self.assertEqual(summary["log_koc_selected"], 1.2)
        self.assertEqual(summary["log_koc_estimated"], 1.05)
        self.assertEqual(summary["log_koc_experimental"], 1.2)

    @patch("src.episuite_io.call_epi_web_api")
    def test_run_epi_web_batch_passes_optional_cas_and_records_raw_traceability(self, call_api):
        call_api.return_value = ETHANOL_CAS_AND_SMILES_RESPONSE
        input_df = pd.DataFrame({"compound": ["Ethanol"], "smiles": ["CCO"], "cas": ["64-17-5"]})

        results, raw_rows, errors = episuite_io.run_epi_web_batch(input_df, delay_seconds=0)

        call_api.assert_called_once_with("CCO", cas="64-17-5", api_url=episuite_io.DEFAULT_EPI_WEB_API, timeout=90)
        self.assertEqual(results.loc[0, "cas"], "64-17-5")
        self.assertEqual(raw_rows.loc[0, "cas"], "64-17-5")
        self.assertEqual(raw_rows.loc[0, "epi_cas"], "000064-17-5")
        self.assertTrue(errors.empty)

    @patch("src.episuite_io.call_epi_web_api")
    def test_run_epi_web_batch_retries_smiles_when_cas_is_not_located(self, call_api):
        call_api.side_effect = [
            RuntimeError("EPI Web Suite 返回 HTTP 404: Could not locate CAS ID, try again with SMILES if available"),
            ETHANOL_CAS_AND_SMILES_RESPONSE,
        ]
        input_df = pd.DataFrame({"compound": ["Ethanol"], "smiles": ["CCO"], "cas": ["64-17-5"]})

        results, raw_rows, errors = episuite_io.run_epi_web_batch(input_df, delay_seconds=0)

        self.assertEqual(call_api.call_args_list[0].kwargs["cas"], "64-17-5")
        self.assertIsNone(call_api.call_args_list[1].kwargs["cas"])
        self.assertEqual(results.loc[0, "status"], "success")
        self.assertIn("CAS 查询失败，已回退到 SMILES", results.loc[0, "query_note"])
        self.assertIn("CAS 查询失败，已回退到 SMILES", raw_rows.loc[0, "query_note"])
        self.assertTrue(errors.empty)

    @patch("src.episuite_io.call_epi_web_api")
    def test_parse_null_with_cas_falls_back_to_same_smiles_without_cas(self, call_api):
        call_api.side_effect = [
            RuntimeError("EPI Web Suite 返回 HTTP 400: could not parse 'null'"),
            ETHANOL_CAS_AND_SMILES_RESPONSE,
        ]
        input_df = pd.DataFrame(
            {
                "compound": ["Example"],
                "smiles": ["CC(C)c1ccc2c(c1)CCC1C(C)CCCC21C"],
                "cas": ["5323-56-8"],
            }
        )

        results, raw_rows, errors = episuite_io.run_epi_web_batch(
            input_df, delay_seconds=0
        )

        self.assertEqual(call_api.call_args_list[0].kwargs["cas"], "5323-56-8")
        self.assertIsNone(call_api.call_args_list[1].kwargs["cas"])
        self.assertEqual(call_api.call_args_list[1].args[0], input_df.loc[0, "smiles"])
        self.assertEqual(results.loc[0, "status"], "success")
        self.assertIn("CAS", results.loc[0, "query_note"])
        self.assertEqual(raw_rows.loc[0, "smiles"], input_df.loc[0, "smiles"])
        self.assertTrue(errors.empty)

    @patch("src.episuite_io.call_epi_web_api")
    def test_other_http_400_does_not_fall_back(self, call_api):
        call_api.side_effect = RuntimeError(
            "EPI Web Suite 返回 HTTP 400: invalid structure"
        )
        input_df = pd.DataFrame(
            {"compound": ["Bad"], "smiles": ["bad"], "cas": ["1-11-1"]}
        )

        results, _, errors = episuite_io.run_epi_web_batch(input_df, delay_seconds=0)

        call_api.assert_called_once()
        self.assertEqual(results.loc[0, "status"], "failed")
        self.assertEqual(len(errors), 1)

    @patch("src.episuite_io.call_epi_web_api")
    @patch("src.episuite_io.resolve_epi_name_exact")
    def test_name_only_row_resolves_and_submits_exact_candidate(self, resolve_name, call_api):
        resolve_name.return_value = {
            "name": "ETHANOL",
            "smiles": "OCC",
            "cas": "000064-17-5",
        }
        call_api.return_value = ETHANOL_CAS_AND_SMILES_RESPONSE

        results, raw_rows, errors = episuite_io.run_epi_web_batch(
            pd.DataFrame({"compound": ["Ethanol"]}),
            delay_seconds=0,
        )

        resolve_name.assert_called_once_with(
            "Ethanol",
            api_url=episuite_io.DEFAULT_EPI_WEB_API,
            timeout=90,
        )
        call_api.assert_called_once_with(
            "OCC",
            cas="000064-17-5",
            api_url=episuite_io.DEFAULT_EPI_WEB_API,
            timeout=90,
        )
        self.assertEqual(results.loc[0, "smiles"], "OCC")
        self.assertEqual(results.loc[0, "cas"], "000064-17-5")
        self.assertIn("名称完全一致", results.loc[0, "query_note"])
        self.assertEqual(raw_rows.loc[0, "smiles"], "OCC")
        self.assertTrue(errors.empty)

    @patch("src.episuite_io.call_epi_web_api")
    @patch("src.episuite_io.resolve_epi_name_exact")
    def test_name_resolution_failure_isolated_without_submit(self, resolve_name, call_api):
        resolve_name.side_effect = RuntimeError("没有名称完全一致的 EPI Suite 候选")

        results, raw_rows, errors = episuite_io.run_epi_web_batch(
            pd.DataFrame({"compound": ["Unknown"]}),
            delay_seconds=0,
        )

        call_api.assert_not_called()
        self.assertTrue(raw_rows.empty)
        self.assertEqual(results.loc[0, "status"], "failed")
        self.assertIn("完全一致", errors.loc[0, "error"])

    @patch("src.episuite_io.call_epi_web_api")
    @patch("src.episuite_io.resolve_epi_name_exact")
    def test_transient_name_resolution_failure_retries_and_succeeds(self, resolve_name, call_api):
        resolve_name.side_effect = [
            RuntimeError("EPI Web Suite 名称搜索返回 HTTP 429: busy"),
            {
                "name": "ETHANOL",
                "smiles": "OCC",
                "cas": "000064-17-5",
            },
        ]
        call_api.return_value = ETHANOL_CAS_AND_SMILES_RESPONSE

        results, raw_rows, errors = episuite_io.run_epi_web_batch(
            pd.DataFrame({"compound": ["Ethanol"]}),
            delay_seconds=0,
        )

        self.assertEqual(resolve_name.call_count, 2)
        self.assertEqual(results.loc[0, "status"], "success")
        self.assertEqual(raw_rows.loc[0, "cas"], "000064-17-5")
        self.assertTrue(errors.empty)

    @patch("src.episuite_io.call_epi_web_api")
    @patch("src.episuite_io.resolve_epi_name_exact")
    def test_resolved_name_cas_fallback_preserves_both_query_notes(
        self, resolve_name, call_api
    ):
        resolve_name.return_value = {
            "name": "ETHANOL",
            "smiles": "OCC",
            "cas": "000064-17-5",
        }
        call_api.side_effect = [
            RuntimeError(
                "EPI Web Suite 返回 HTTP 404: Could not locate CAS ID, "
                "try again with SMILES if available"
            ),
            ETHANOL_CAS_AND_SMILES_RESPONSE,
        ]

        results, raw_rows, errors = episuite_io.run_epi_web_batch(
            pd.DataFrame({"compound": ["Ethanol"]}),
            delay_seconds=0,
        )

        self.assertEqual(call_api.call_args_list[0].kwargs["cas"], "000064-17-5")
        self.assertIsNone(call_api.call_args_list[1].kwargs["cas"])
        for note in (results.loc[0, "query_note"], raw_rows.loc[0, "query_note"]):
            self.assertIn("名称完全一致", note)
            self.assertIn("CAS 查询失败，已回退到 SMILES", note)
        self.assertEqual(raw_rows.loc[0, "smiles"], "OCC")
        self.assertEqual(raw_rows.loc[0, "cas"], "000064-17-5")
        self.assertTrue(errors.empty)

    @patch("src.episuite_io.call_epi_web_api")
    @patch("src.episuite_io.resolve_epi_name_exact")
    def test_mixed_name_failures_keep_order_and_isolate_rows(self, resolve_name, call_api):
        def resolve(compound, **kwargs):
            if compound == "Unknown":
                raise RuntimeError("没有名称完全一致的 EPI Suite 候选")
            return {
                "name": "ETHANOL",
                "smiles": "OCC",
                "cas": "000064-17-5",
            }

        resolve_name.side_effect = resolve
        call_api.return_value = ETHANOL_CAS_AND_SMILES_RESPONSE
        input_df = pd.DataFrame(
            {
                "compound": ["Ethanol", "Unknown", "Direct SMILES"],
                "smiles": [pd.NA, pd.NA, "CCC"],
                "cas": [pd.NA, pd.NA, "3-33-3"],
            }
        )

        results, raw_rows, errors = episuite_io.run_epi_web_batch(
            input_df,
            delay_seconds=0,
            max_workers=3,
        )

        self.assertEqual(
            results["compound"].tolist(),
            ["Ethanol", "Unknown", "Direct SMILES"],
        )
        self.assertEqual(results["status"].tolist(), ["success", "failed", "success"])
        self.assertEqual(errors["compound"].tolist(), ["Unknown"])
        self.assertEqual(raw_rows["compound"].tolist(), ["Ethanol", "Direct SMILES"])
        self.assertEqual(raw_rows.loc[0, "smiles"], "OCC")
        self.assertEqual(raw_rows.loc[0, "cas"], "000064-17-5")

    @patch("src.episuite_io.call_epi_web_api")
    def test_run_epi_web_batch_keeps_order_and_isolates_parallel_row_failure(self, call_api):
        def fake_call(smiles, **kwargs):
            if smiles == "bad":
                raise RuntimeError("network failed")
            raw = dict(ETHANOL_CAS_AND_SMILES_RESPONSE)
            raw["parameters"] = {"smiles": smiles}
            return raw

        call_api.side_effect = fake_call
        input_df = pd.DataFrame(
            {
                "compound": ["A", "B", "C", "D", "E"],
                "smiles": ["a", "b", "bad", "d", "e"],
            }
        )

        results, raw_rows, errors = episuite_io.run_epi_web_batch(
            input_df,
            delay_seconds=0,
            max_workers=3,
        )

        self.assertEqual(results["compound"].tolist(), ["A", "B", "C", "D", "E"])
        self.assertEqual(results.loc[2, "status"], "failed")
        self.assertEqual(errors.loc[0, "compound"], "C")
        self.assertEqual(len(raw_rows), 4)

    @patch("src.episuite_io.call_epi_web_api")
    def test_build_epi_web_result_tables_splits_raw_json_into_category_tables(self, call_api):
        call_api.return_value = ETHANOL_CAS_AND_SMILES_RESPONSE
        input_df = pd.DataFrame({"compound": ["Ethanol"], "smiles": ["CCO"], "cas": ["64-17-5"]})
        core, raw_rows, errors = episuite_io.run_epi_web_batch(input_df, delay_seconds=0)

        tables = episuite_io.build_epi_web_result_tables(core, raw_rows, errors)

        expected = {
            "Core_Summary",
            "Properties",
            "Degradation",
            "Fate_Transport",
            "Bioaccumulation",
            "ECOSAR_Aquatic_Toxicity",
            "Model_Metadata",
            "Raw_API_JSON",
            "Warnings",
        }
        self.assertEqual(set(tables), expected)
        core_columns = set(tables["Core_Summary"].columns)
        self.assertNotIn("log_kow", core_columns)
        self.assertNotIn("log_kow_selected", core_columns)
        self.assertEqual(tables["Core_Summary"].loc[0, "log_kow_estimated"], -0.1411999762058258)
        self.assertEqual(tables["Core_Summary"].loc[0, "log_kow_experimental"], -0.31)
        self.assertEqual(tables["Core_Summary"].loc[0, "log_kow_type"], "EXPERIMENTAL")

        property_columns = set(tables["Properties"].columns)
        self.assertNotIn("log_kow_selected", property_columns)
        self.assertNotIn("log_kow_units", property_columns)
        self.assertEqual(tables["Properties"].loc[0, "log_kow_estimated"], -0.1411999762058258)
        self.assertEqual(tables["Properties"].loc[0, "log_kow_experimental"], -0.31)
        self.assertEqual(tables["Properties"].loc[0, "log_kow_type"], "EXPERIMENTAL")
        self.assertEqual(tables["Degradation"].loc[0, "biowin_Ultimate_Biodegradation_Timeframe"], 2.1)
        self.assertEqual(tables["Fate_Transport"].loc[0, "level3_water_percent"], 70.0)
        self.assertEqual(tables["Bioaccumulation"].loc[0, "bcf"], 3.2)
        self.assertEqual(len(tables["ECOSAR_Aquatic_Toxicity"]), 2)
        self.assertEqual(tables["ECOSAR_Aquatic_Toxicity"].loc[0, "endpoint"], "LC50")
        self.assertIn("raw_json", tables["Raw_API_JSON"].columns)

    def test_properties_include_partition_pairs_and_rdkit_descriptors(self):
        response = self._response_with_koawin_model()
        raw_rows = pd.DataFrame(
            [
                {
                    "compound": "Ethanol",
                    "smiles": "CC",
                    "cas": "64-17-5",
                    "epi_smiles": "c1ccccc1",
                    "raw_json": json.dumps(response),
                }
            ]
        )

        tables = episuite_io.build_epi_web_result_tables(raw_df=raw_rows)
        properties = tables["Properties"]

        expected_columns = [
            "koawin_log_kow",
            "koawin_kow",
            "koawin_log_koa",
            "koawin_koa",
            "koawin_log_kaw",
            "koawin_kaw",
            "tpsa_rdkit_a2",
            "mr_rdkit_cm3_mol",
        ]
        positions = [properties.columns.get_loc(name) for name in expected_columns]
        self.assertEqual(positions, list(range(positions[0], positions[0] + 8)))
        for column in expected_columns:
            self.assertIn(column, properties.columns)
            for sheet_name in episuite_io.EPI_WEB_RESULT_SHEETS:
                if sheet_name != "Properties":
                    self.assertNotIn(
                        column,
                        tables[sheet_name].columns,
                        msg=f"{column} leaked into {sheet_name}",
                    )
        self.assertAlmostEqual(properties.loc[0, "tpsa_rdkit_a2"], 20.23)
        self.assertAlmostEqual(properties.loc[0, "mr_rdkit_cm3_mol"], 12.7598)
        self.assertTrue(
            math.isclose(
                properties.loc[0, "koawin_log_kaw"],
                math.log10(properties.loc[0, "koawin_kaw"]),
            )
        )
        self.assertEqual(
            json.loads(tables["Raw_API_JSON"].loc[0, "raw_json"]),
            response,
        )
        self.assertTrue(tables["Warnings"].empty)

    def test_descriptor_failure_is_added_to_warnings_without_dropping_properties(self):
        response = self._response_with_koawin_model()
        response["chemicalProperties"]["smiles"] = "not-a-smiles"
        raw_rows = pd.DataFrame(
            [
                {
                    "compound": "Broken structure",
                    "smiles": "",
                    "cas": "",
                    "epi_smiles": "",
                    "raw_json": json.dumps(response),
                }
            ]
        )
        existing_warning = {
            "compound": "Existing compound",
            "smiles": "CCC",
            "cas": "111-11-1",
            "warning": "已有警告",
        }

        tables = episuite_io.build_epi_web_result_tables(
            raw_df=raw_rows,
            warnings_df=pd.DataFrame([existing_warning]),
        )

        self.assertEqual(len(tables["Properties"]), 1)
        for column in ("tpsa_rdkit_a2", "mr_rdkit_cm3_mol"):
            self.assertTrue(pd.isna(tables["Properties"].loc[0, column]))

        warnings = tables["Warnings"]
        retained = warnings.loc[warnings["warning"].eq(existing_warning["warning"])]
        self.assertEqual(
            retained[["compound", "smiles", "cas", "warning"]].to_dict("records"),
            [existing_warning],
        )
        descriptor_warning = warnings.loc[
            warnings["warning"].eq("RDKit 描述符未计算：SMILES 无法解析")
        ]
        self.assertEqual(
            descriptor_warning[
                ["compound", "smiles", "cas", "warning"]
            ].to_dict("records"),
            [
                {
                    "compound": "Broken structure",
                    "smiles": "",
                    "cas": "",
                    "warning": "RDKit 描述符未计算：SMILES 无法解析",
                }
            ],
        )

    def test_koawin_inconsistency_warning_keeps_identity_and_existing_warnings(self):
        response = self._response_with_koawin_model()
        response["logKoa"]["estimatedValue"]["model"]["koa"] *= 2
        raw_rows = pd.DataFrame(
            [
                {
                    "compound": "Inconsistent partition",
                    "smiles": "CCO",
                    "cas": "64-17-5",
                    "raw_json": json.dumps(response),
                }
            ]
        )
        existing_warning = {
            "compound": "Existing compound",
            "smiles": "CCC",
            "cas": "111-11-1",
            "warning": "已有警告",
        }

        tables = episuite_io.build_epi_web_result_tables(
            raw_df=raw_rows,
            warnings_df=pd.DataFrame([existing_warning]),
        )

        warnings = tables["Warnings"]
        retained = warnings.loc[warnings["warning"].eq(existing_warning["warning"])]
        self.assertEqual(
            retained[["compound", "smiles", "cas", "warning"]].to_dict("records"),
            [existing_warning],
        )
        coefficient_warning = warnings.loc[
            warnings["warning"].eq(
                "KOAWIN 原始系数关系不一致：KOA != KOW / KAW"
            )
        ]
        self.assertEqual(
            coefficient_warning[
                ["compound", "smiles", "cas", "warning"]
            ].to_dict("records"),
            [
                {
                    "compound": "Inconsistent partition",
                    "smiles": "CCO",
                    "cas": "64-17-5",
                    "warning": "KOAWIN 原始系数关系不一致：KOA != KOW / KAW",
                }
            ],
        )

    @patch("src.episuite_io.call_epi_web_api")
    def test_result_workbook_writes_category_tables_and_raw_json(self, call_api):
        call_api.return_value = ETHANOL_CAS_AND_SMILES_RESPONSE
        input_df = pd.DataFrame({"compound": ["Ethanol"], "smiles": ["CCO"], "cas": ["64-17-5"]})
        core, raw_rows, errors = episuite_io.run_epi_web_batch(input_df, delay_seconds=0)

        workbook = episuite_io.build_result_workbook(
            input_df,
            merged_df=core,
            parsed_df=core,
            warnings_df=errors,
            raw_df=raw_rows,
        )
        sheets = set(load_workbook(workbook, read_only=True).sheetnames)

        self.assertIn("Core_Summary", sheets)
        self.assertIn("ECOSAR_Aquatic_Toxicity", sheets)
        self.assertIn("Raw_API_JSON", sheets)


if __name__ == "__main__":
    unittest.main()
