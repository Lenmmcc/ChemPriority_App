import io
import unittest
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

from src.source_origin import build_result_workbook, fetch_coconut_evidence, run_source_origin_batch


def _input_row(compound="Bisphenol A"):
    return {
        "compound": compound,
        "cas": "80-05-7",
        "ec": "201-245-8",
        "smiles": "CC(C)(c1ccc(O)cc1)c1ccc(O)cc1",
        "dtxsid": "DTXSID7020182",
        "echa_id": "100.001.133",
    }


def _comptox_candidate(compound="Bisphenol A"):
    return {
        "compound": compound,
        "dtxsid": "DTXSID7020182",
        "source_type": "product_category",
        "source": "CompTox Dashboard",
        "use_cn": "塑料/聚合物制品",
        "raw_use": "Plastic products",
        "evidence_count": 4,
    }


def _echa_candidate(compound="Bisphenol A"):
    return {
        "compound": compound,
        "echa_id": "100.001.133",
        "use_cn": "工业用途",
        "use_en": "Industrial use",
        "raw_use": "Industrial use",
        "evidence_count": 1,
        "source": "ECHA REACH dossier",
        "record_url": "https://chem.echa.europa.eu/html-pages-prod/example/index.html",
    }


def _natural_evidence(source_name="ChEBI", confidence="strong"):
    return {
        "source_group": "natural",
        "source_name": source_name,
        "evidence_type": "natural_product_match",
        "evidence_label": "naturally occurring metabolite",
        "evidence_text": "A naturally occurring metabolite recorded for this compound.",
        "matched_identifier": "CHEBI:123",
        "confidence": confidence,
        "record_url": "https://www.ebi.ac.uk/chebi/searchId.do?chebiId=CHEBI:123",
    }


class SourceOriginBatchTests(unittest.TestCase):
    @patch("src.source_origin.fetch_coconut_evidence", return_value=[])
    @patch("src.source_origin.fetch_chebi_evidence", return_value=[])
    def test_human_evidence_only_is_human_source(self, fetch_chebi, fetch_coconut):
        summary, evidence, errors = run_source_origin_batch(
            pd.DataFrame([_input_row()]),
            comptox_candidates_df=pd.DataFrame([_comptox_candidate()]),
            echa_candidates_df=pd.DataFrame([_echa_candidate()]),
            delay_seconds=0,
        )

        self.assertTrue(errors.empty)
        self.assertEqual(summary.loc[0, "来源属性"], "人为源")
        self.assertEqual(summary.loc[0, "人为源证据数"], 2)
        self.assertEqual(summary.loc[0, "天然源证据数"], 0)
        self.assertIn("EPA CompTox", set(evidence["source_name"]))
        self.assertIn("ECHA CHEM", set(evidence["source_name"]))
        self.assertTrue(evidence["source_group"].eq("human").all())

    @patch("src.source_origin.fetch_coconut_evidence", return_value=[])
    @patch("src.source_origin.fetch_chebi_evidence", return_value=[_natural_evidence()])
    def test_natural_evidence_only_is_natural_source(self, fetch_chebi, fetch_coconut):
        summary, evidence, errors = run_source_origin_batch(
            pd.DataFrame([_input_row("Caffeine")]),
            comptox_candidates_df=pd.DataFrame(),
            echa_candidates_df=pd.DataFrame(),
            delay_seconds=0,
        )

        self.assertTrue(errors.empty)
        self.assertEqual(summary.loc[0, "来源属性"], "天然源")
        self.assertEqual(summary.loc[0, "人为源证据数"], 0)
        self.assertEqual(summary.loc[0, "天然源证据数"], 1)
        self.assertEqual(evidence.loc[0, "source_name"], "ChEBI")

    @patch("src.source_origin.fetch_coconut_evidence", return_value=[_natural_evidence("COCONUT")])
    @patch("src.source_origin.fetch_chebi_evidence", return_value=[])
    def test_human_and_natural_evidence_is_both_source(self, fetch_chebi, fetch_coconut):
        summary, evidence, errors = run_source_origin_batch(
            pd.DataFrame([_input_row("Caffeine")]),
            comptox_candidates_df=pd.DataFrame([_comptox_candidate("Caffeine")]),
            echa_candidates_df=pd.DataFrame(),
            delay_seconds=0,
        )

        self.assertTrue(errors.empty)
        self.assertEqual(summary.loc[0, "来源属性"], "兼具天然源和人为源")
        self.assertEqual(summary.loc[0, "人为源证据数"], 1)
        self.assertEqual(summary.loc[0, "天然源证据数"], 1)
        self.assertEqual(set(evidence["source_group"]), {"human", "natural"})

    @patch("src.source_origin.fetch_coconut_evidence", return_value=[])
    @patch("src.source_origin.fetch_chebi_evidence", return_value=[_natural_evidence(confidence="weak")])
    def test_weak_or_empty_evidence_is_insufficient(self, fetch_chebi, fetch_coconut):
        summary, evidence, errors = run_source_origin_batch(
            pd.DataFrame([_input_row("Unknown")]),
            comptox_candidates_df=pd.DataFrame(),
            echa_candidates_df=pd.DataFrame(),
            delay_seconds=0,
        )

        self.assertTrue(errors.empty)
        self.assertEqual(summary.loc[0, "来源属性"], "证据不足")
        self.assertEqual(summary.loc[0, "证据等级"], "弱")
        self.assertEqual(summary.loc[0, "天然源证据数"], 0)
        self.assertEqual(evidence.loc[0, "confidence"], "weak")

    @patch("src.source_origin.fetch_coconut_evidence", return_value=[])
    @patch("src.source_origin.fetch_chebi_evidence", side_effect=RuntimeError("timeout"))
    def test_external_failure_keeps_summary_row_and_warning(self, fetch_chebi, fetch_coconut):
        summary, evidence, errors = run_source_origin_batch(
            pd.DataFrame([_input_row("Caffeine")]),
            comptox_candidates_df=pd.DataFrame(),
            echa_candidates_df=pd.DataFrame(),
            delay_seconds=0,
        )

        self.assertTrue(evidence.empty)
        self.assertEqual(summary.loc[0, "compound"], "Caffeine")
        self.assertEqual(summary.loc[0, "来源属性"], "证据不足")
        self.assertEqual(errors.loc[0, "source_name"], "ChEBI")
        self.assertIn("timeout", errors.loc[0, "message"])

    @patch("src.source_origin.fetch_coconut_evidence", return_value=[])
    @patch("src.source_origin.fetch_chebi_evidence", return_value=[_natural_evidence()])
    def test_result_workbook_contains_summary_evidence_and_warnings(self, fetch_chebi, fetch_coconut):
        summary, evidence, errors = run_source_origin_batch(
            pd.DataFrame([_input_row("Caffeine")]),
            comptox_candidates_df=pd.DataFrame(),
            echa_candidates_df=pd.DataFrame(),
            delay_seconds=0,
        )

        workbook = build_result_workbook(
            pd.DataFrame([_input_row("Caffeine")]),
            summary_df=summary,
            evidence_df=evidence,
            errors_df=errors,
        )
        book = load_workbook(io.BytesIO(workbook.getvalue()), read_only=True)

        self.assertIn("Source_Origin_Summary", book.sheetnames)
        self.assertIn("Source_Origin_Evidence", book.sheetnames)
        self.assertIn("Source_Origin_Warnings", book.sheetnames)

    @patch("src.source_origin.fetch_coconut_evidence", return_value=[])
    @patch("src.source_origin.fetch_chebi_evidence", return_value=[])
    def test_predicted_comptox_functional_use_is_medium_confidence(self, fetch_chebi, fetch_coconut):
        predicted_candidate = {
            "compound": "Example",
            "dtxsid": "DTXSID0000001",
            "source_type": "functional_use",
            "source": "dashboard:functional_use",
            "use_cn": "芳香剂",
            "raw_use": "fragrance",
            "probability": 0.91,
            "functional_use_source": "predicted",
        }

        _, evidence, _ = run_source_origin_batch(
            pd.DataFrame([{**_input_row("Example"), "dtxsid": "DTXSID0000001"}]),
            comptox_candidates_df=pd.DataFrame([predicted_candidate]),
            echa_candidates_df=pd.DataFrame(),
            delay_seconds=0,
        )

        self.assertEqual(evidence.loc[0, "evidence_type"], "functional_use")
        self.assertEqual(evidence.loc[0, "confidence"], "medium")

    @patch("src.source_origin.time.sleep")
    @patch("src.source_origin._post_json")
    def test_coconut_lookup_retries_transient_post_failure(self, post_json, sleep):
        post_json.side_effect = [
            RuntimeError("Remote end closed connection without response"),
            {
                "data": {
                    "data": [
                        {
                            "identifier": "CNP0228556.0",
                            "canonical_smiles": "CN1C(=O)C2=C(N=CN2C)N(C)C1=O",
                            "name": "caffeine",
                            "organism_count": 135,
                            "citation_count": 13,
                            "collection_count": 31,
                        }
                    ]
                }
            },
        ]

        evidence = fetch_coconut_evidence(pd.Series({"compound": "caffeine"}), timeout=1)

        self.assertEqual(evidence[0]["source_name"], "COCONUT")
        self.assertEqual(evidence[0]["confidence"], "strong")
        self.assertEqual(post_json.call_count, 2)
        sleep.assert_called_once_with(1.0)

    @patch("src.source_origin._post_json")
    def test_coconut_lookup_rejects_unmatched_search_hit(self, post_json):
        post_json.return_value = {
            "data": {
                "data": [
                    {
                        "identifier": "CNP0511569.0",
                        "canonical_smiles": "CCBr",
                        "name": "Bromoethane",
                        "organism_count": 0,
                        "citation_count": 0,
                        "collection_count": 1,
                    }
                ]
            }
        }

        evidence = fetch_coconut_evidence(
            pd.Series(
                {
                    "compound": "3cH2B",
                    "cas": "84540-37-4",
                    "smiles": "CCCC1CCC(CC1)C2=CC=C(C=C2)C3=CC=C(C=C3)CC",
                }
            ),
            timeout=1,
        )

        self.assertEqual(evidence, [])


if __name__ == "__main__":
    unittest.main()
