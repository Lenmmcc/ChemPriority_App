import io
import ast
import unittest
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook

import src.cp_screening_workflow as workflow
from src.mol_structure_parser import find_mol_text_column, prepare_structure_dataframe
from src.cp_screening_workflow import (
    EXPECTED_WORKBOOK_SHEETS,
    PBMToxPiConfig,
    PBMToxPiResult,
    build_detection_frequency,
    build_peak_area_long,
    build_pbm_toxpi_input,
    build_screening_workbook,
    calculate_pbm_toxpi,
    generate_pbm_toxpi_bar_plot,
    generate_pbm_toxpi_robustness_plot,
    limit_toxpi_plot_rows,
    run_pbm_toxpi_robustness,
)


ETHANOL_MOL = """ethanol
  ChemPriority

  3  2  0  0  0  0  0  0  0  0  0
    0.0000    0.0000    0.0000 C   0  0  0  0  0  0  0  0  0  0  0  0
    1.5000    0.0000    0.0000 C   0  0  0  0  0  0  0  0  0  0  0  0
    2.2500    1.2990    0.0000 O   0  0  0  0  0  0  0  0  0  0  0  0
  1  2  1  0
  2  3  1  0
M  END
"""


def load_screening_mapping_normalizer():
    page_path = Path("pages/0_综合筛查流程.py")
    page_module = ast.parse(page_path.read_text(encoding="utf-8"), filename=str(page_path))
    function = next(
        node
        for node in page_module.body
        if isinstance(node, ast.FunctionDef) and node.name == "normalize_samples_for_mappings"
    )
    namespace = {
        "pd": pd,
        "prepare_structure_dataframe": prepare_structure_dataframe,
        "STANDARD_COMPOUND_COL": "Name",
        "STANDARD_FORMULA_COL": "formula",
        "STANDARD_SMILES_COL": "SMILES_input",
        "STANDARD_CAS_COL": "CAS_input",
        "clean_text": lambda value: "" if value is None or pd.isna(value) else str(value).strip(),
    }
    exec(compile(ast.Module(body=[function], type_ignores=[]), str(page_path), "exec"), namespace)
    return namespace["normalize_samples_for_mappings"]


class CpScreeningWorkflowTests(unittest.TestCase):
    def test_pbm_toxpi_bar_plot_uses_times_new_roman(self):
        toxpi_results = pd.DataFrame(
            {
                "compound": ["Compound A", "Compound B"],
                "toxpi": [0.8, 0.4],
            }
        )

        figure = generate_pbm_toxpi_bar_plot(toxpi_results)
        try:
            texts = [
                text
                for text in figure.findobj(matplotlib.text.Text)
                if text.get_text().strip()
            ]
            self.assertTrue(texts)
            self.assertTrue(
                all(text.get_fontfamily()[0] == "Times New Roman" for text in texts)
            )
        finally:
            plt.close(figure)

    def test_screening_default_mapping_detects_recognized_mol_column(self):
        page_path = Path("pages/0_综合筛查流程.py")
        page_module = ast.parse(page_path.read_text(encoding="utf-8"), filename=str(page_path))
        function = next(
            node
            for node in page_module.body
            if isinstance(node, ast.FunctionDef) and node.name == "sample_mapping_defaults"
        )
        namespace = {
            "guess_column": lambda columns, candidates, fallback_index=0: columns[fallback_index],
            "guess_peak_area_column": lambda columns: columns[0],
            "group_area_columns": lambda columns: [],
            "find_mol_text_column": find_mol_text_column,
        }
        exec(compile(ast.Module(body=[function], type_ignores=[]), str(page_path), "exec"), namespace)

        defaults = namespace["sample_mapping_defaults"](
            {"data": pd.DataFrame({"Name": ["Ethanol"], " Structure ": [ETHANOL_MOL]})}
        )

        self.assertEqual(defaults["mol_column"], " Structure ")

    def test_screening_mapping_uses_mol_derived_smiles_for_downstream_input(self):
        normalize_samples = load_screening_mapping_normalizer()
        samples = [
            {
                "name": "Ethanol",
                "data": pd.DataFrame(
                    {
                        "Compound": ["Ethanol"],
                        "Formula": ["C2H6O"],
                        "Peak_Area": [100.0],
                        "Structure": [ETHANOL_MOL],
                    }
                ),
            }
        ]
        mappings = {
            "Ethanol": {
                "compound_col": "Compound",
                "formula_col": "Formula",
                "peak_area_col": "Peak_Area",
                "sample_cols": ["Peak_Area"],
                "mol_column": "Structure",
                "smiles_col": None,
                "cas_col": None,
            }
        }

        normalized, _, _ = normalize_samples(samples, mappings)

        self.assertIn("SMILES_input", normalized[0]["data"].columns)
        self.assertEqual(normalized[0]["data"].loc[0, "SMILES_input"], "CCO")

    def test_screening_mapping_preserves_raw_smiles_when_mol_is_the_fallback(self):
        normalize_samples = load_screening_mapping_normalizer()
        samples = [
            {
                "name": "Ethanol",
                "data": pd.DataFrame(
                    {
                        "Compound": ["Ethanol"],
                        "Formula": ["C2H6O"],
                        "Peak_Area": [100.0],
                        "Structure": [ETHANOL_MOL],
                        "smiles": ["not valid"],
                    }
                ),
            }
        ]
        mappings = {
            "Ethanol": {
                "compound_col": "Compound",
                "formula_col": "Formula",
                "peak_area_col": "Peak_Area",
                "sample_cols": ["Peak_Area"],
                "mol_column": "Structure",
                "smiles_col": "smiles",
                "cas_col": None,
            }
        }

        normalized, _, _ = normalize_samples(samples, mappings)

        self.assertEqual(normalized[0]["data"].loc[0, "smiles"], "not valid")
        self.assertEqual(normalized[0]["data"].loc[0, "SMILES_input"], "CCO")

    def test_detection_frequency_uses_uploaded_files_as_samples(self):
        samples = [
            (
                "S1.xlsx",
                pd.DataFrame(
                    {
                        "Name": ["Compound A", "Compound B", "Compound A"],
                        "Group_Area": [100001, 50000, 200000],
                    }
                ),
            ),
            (
                "S2.xlsx",
                pd.DataFrame(
                    {
                        "Name": ["Compound A", "Compound C"],
                        "Group_Area": [100000, 300000],
                    }
                ),
            ),
            (
                "S3.xlsx",
                pd.DataFrame(
                    {
                        "Name": ["Compound B", "Compound A"],
                        "Group_Area": [110000, np.nan],
                    }
                ),
            ),
        ]

        df_table, sample_peak_area = build_detection_frequency(samples, detection_threshold=1e5)

        by_compound = df_table.set_index("compound")
        self.assertEqual(by_compound.loc["Compound A", "detected_sample_count"], 1)
        self.assertEqual(by_compound.loc["Compound A", "total_sample_count"], 3)
        self.assertAlmostEqual(by_compound.loc["Compound A", "DF"], 1 / 3)
        self.assertEqual(by_compound.loc["Compound B", "detected_sample_count"], 1)
        self.assertEqual(by_compound.loc["Compound C", "detected_sample_count"], 1)

        a_s1 = sample_peak_area[
            (sample_peak_area["compound"].eq("Compound A"))
            & (sample_peak_area["sample_id"].eq("S1"))
        ].iloc[0]
        self.assertEqual(a_s1["peak_area"], 200000)
        self.assertTrue(bool(a_s1["detected"]))

        a_s2 = sample_peak_area[
            (sample_peak_area["compound"].eq("Compound A"))
            & (sample_peak_area["sample_id"].eq("S2"))
        ].iloc[0]
        self.assertFalse(bool(a_s2["detected"]))

    def test_detection_frequency_uses_file_level_group_area_mean(self):
        samples = [
            (
                "S1.xlsx",
                pd.DataFrame(
                    {
                        "Name": ["Compound A", "Compound B"],
                        "Group Area: S1-1": [100000, 50000],
                        "Group Area: S1-2": [100001, 250000],
                    }
                ),
            ),
            (
                "S2.xlsx",
                pd.DataFrame(
                    {
                        "Name": ["Compound A", "Compound B"],
                        "Group Area: S1-1": [90000, 100000],
                        "Group Area: S1-2": [80000, np.nan],
                    }
                ),
            ),
        ]

        df_table, sample_peak_area = build_detection_frequency(
            samples,
            peak_area_col=["Group Area: S1-1", "Group Area: S1-2"],
            detection_threshold=1e5,
        )

        by_compound = df_table.set_index("compound")
        self.assertEqual(by_compound.loc["Compound A", "detected_sample_count"], 1)
        self.assertEqual(by_compound.loc["Compound B", "detected_sample_count"], 1)
        self.assertEqual(by_compound.loc["Compound A", "total_sample_count"], 2)
        self.assertAlmostEqual(by_compound.loc["Compound A", "DF"], 0.5)

        a_s1 = sample_peak_area[
            (sample_peak_area["compound"].eq("Compound A"))
            & (sample_peak_area["source_sample_id"].eq("S1"))
        ].iloc[0]
        self.assertEqual(a_s1["sample_point"], "Group_Area_Mean")
        self.assertAlmostEqual(a_s1["peak_area"], 100000.5)
        self.assertTrue(bool(a_s1["detected"]))

        b_s2 = sample_peak_area[
            (sample_peak_area["compound"].eq("Compound B"))
            & (sample_peak_area["source_sample_id"].eq("S2"))
        ].iloc[0]
        self.assertEqual(b_s2["sample_point"], "Group_Area_Mean")
        self.assertAlmostEqual(b_s2["peak_area"], 100000)
        self.assertFalse(bool(b_s2["detected"]))
        self.assertEqual(len(sample_peak_area), 4)

    def test_detection_frequency_allows_distinct_group_area_columns_per_file(self):
        samples = [
            (
                "HH-Aci.xlsx",
                pd.DataFrame(
                    {
                        "Name": ["Compound A", "Compound B"],
                        "Group Area: HH-ACI": [200000, 50000],
                    }
                ),
            ),
            (
                "WH-Aci.xlsx",
                pd.DataFrame(
                    {
                        "Name": ["Compound A", "Compound B"],
                        "Group Area: WH-ACI": [90000, 300000],
                    }
                ),
            ),
            (
                "YC-zhong-Level3.xlsx",
                pd.DataFrame(
                    {
                        "Name": ["Compound A", "Compound B"],
                        "Group Area: 01YiChang-zhong-1": [120000, 100000],
                        "Group Area: 01YiChang-zhong-2": [80000, 300000],
                    }
                ),
            ),
        ]

        selected_cols = [
            "Group Area: HH-ACI",
            "Group Area: WH-ACI",
            "Group Area: 01YiChang-zhong-1",
            "Group Area: 01YiChang-zhong-2",
        ]
        df_table, sample_peak_area = build_detection_frequency(
            samples,
            peak_area_col=selected_cols,
            detection_threshold=1e5,
        )
        mean_df = workflow.build_group_area_mean_by_sample(
            samples,
            peak_area_cols=selected_cols,
        )
        raw_long = build_peak_area_long(samples, peak_area_cols=selected_cols)

        by_compound = df_table.set_index("compound")
        self.assertEqual(by_compound.loc["Compound A", "total_sample_count"], 3)
        self.assertEqual(by_compound.loc["Compound A", "detected_sample_count"], 1)
        self.assertAlmostEqual(by_compound.loc["Compound A", "DF"], 1 / 3)
        self.assertEqual(by_compound.loc["Compound B", "detected_sample_count"], 2)
        self.assertAlmostEqual(by_compound.loc["Compound B", "DF"], 2 / 3)
        self.assertEqual(len(sample_peak_area), 6)
        self.assertEqual(len(mean_df), 6)
        self.assertEqual(len(raw_long), 8)

        yc_a = mean_df[
            mean_df["source_sample_id"].eq("YC-zhong-Level3")
            & mean_df["compound"].eq("Compound A")
        ].iloc[0]
        self.assertAlmostEqual(yc_a["Group_Area_Mean"], 100000)

    def test_detection_frequency_skips_files_without_selected_group_area_columns(self):
        samples = [
            (
                "Included.xlsx",
                pd.DataFrame(
                    {
                        "Name": ["Compound A"],
                        "Group Area: Included": [200000],
                    }
                ),
            ),
            (
                "Skipped.xlsx",
                pd.DataFrame(
                    {
                        "Name": ["Compound A"],
                        "Group Area: Other": [300000],
                    }
                ),
            ),
        ]

        df_table, sample_peak_area = build_detection_frequency(
            samples,
            peak_area_col=["Group Area: Included"],
            detection_threshold=1e5,
        )

        self.assertEqual(df_table.loc[0, "total_sample_count"], 1)
        self.assertEqual(df_table.loc[0, "detected_sample_count"], 1)
        self.assertEqual(set(sample_peak_area["source_sample_id"]), {"Included"})

    def test_group_area_mean_ignores_blank_values_and_keeps_zero_values(self):
        helper = getattr(workflow, "build_group_area_mean_by_sample", None)
        self.assertIsNotNone(helper, "build_group_area_mean_by_sample should be implemented")
        samples = [
            (
                "Sample.xlsx",
                pd.DataFrame(
                    {
                        "Name": ["Compound A", "Compound B"],
                        "formula": ["C10H20", "C12H24"],
                        "Group Area: P1": [np.nan, 10.0],
                        "Group Area: P2": [0.0, np.nan],
                        "Group Area: P3": [100.0, 30.0],
                    }
                ),
            )
        ]

        mean_df = helper(
            samples,
            compound_col="Name",
            formula_col="formula",
            peak_area_cols=["Group Area: P1", "Group Area: P2", "Group Area: P3"],
        )

        by_compound = mean_df.set_index("compound")
        self.assertAlmostEqual(by_compound.loc["Compound A", "Group_Area_Mean"], 50.0)
        self.assertAlmostEqual(by_compound.loc["Compound A", "Peak_Area"], 50.0)
        self.assertAlmostEqual(by_compound.loc["Compound A", "ir_value"], np.log10(50.0))
        self.assertEqual(by_compound.loc["Compound A", "group_area_count"], 2)
        self.assertAlmostEqual(by_compound.loc["Compound B", "Group_Area_Mean"], 20.0)
        self.assertEqual(by_compound.loc["Compound B", "group_area_columns"], "Group Area: P1; Group Area: P2; Group Area: P3")

    def test_peak_area_long_preserves_compound_by_sample_point_rows(self):
        samples = [
            (
                "SampleFile",
                pd.DataFrame(
                    {
                        "Name": ["Compound A", "Compound B"],
                        "formula": ["C10H20", "C12H24"],
                        "Group Area: P1": [100.0, 0.0],
                        "Group Area: P2": [10000.0, 1000.0],
                    }
                ),
            )
        ]

        long_df = build_peak_area_long(
            samples,
            compound_col="Name",
            formula_col="formula",
            peak_area_cols=["Group Area: P1", "Group Area: P2"],
        )

        self.assertEqual(len(long_df), 4)
        self.assertEqual(set(long_df["source_sample_id"]), {"SampleFile"})
        self.assertEqual(set(long_df["sample_id"]), {"Group Area: P1", "Group Area: P2"})
        a_p2 = long_df[
            (long_df["compound"].eq("Compound A"))
            & (long_df["sample_id"].eq("Group Area: P2"))
        ].iloc[0]
        self.assertEqual(a_p2["Peak_Area"], 10000.0)
        self.assertAlmostEqual(a_p2["ir_value"], 4.0)

    def test_pbm_toxpi_input_uses_file_level_mean_rows_before_compound_average(self):
        df_table = pd.DataFrame(
            {
                "compound": ["Compound A", "Compound B"],
                "compound_key": ["compound a", "compound b"],
                "DF": [0.5, 1.0],
                "Peak_Area": [10000.0, 1000.0],
            }
        )
        pov_lrtp = pd.DataFrame({"Name": ["Compound A", "Compound B"], "Scores": [2.0, 4.0]})
        group_area_mean = pd.DataFrame(
            {
                "source_sample_id": ["File1", "File1", "File2"],
                "sample_id": ["File1", "File1", "File2"],
                "compound": ["Compound A", "Compound A", "Compound B"],
                "Peak_Area": [5050.0, 12000.0, 1000.0],
            }
        )

        toxpi_input = build_pbm_toxpi_input(df_table, pov_lrtp, peak_area_long=group_area_mean)
        result = calculate_pbm_toxpi(toxpi_input)
        normalized = result.candidate_normalized
        toxpi_results = result.final_ranking

        self.assertEqual(len(toxpi_input), 3)
        self.assertEqual(toxpi_input["compound"].tolist().count("Compound A"), 2)
        self.assertEqual(set(toxpi_input["sample_id"].dropna()), {"File1", "File2"})

        by_compound = toxpi_results.set_index("compound")
        self.assertAlmostEqual(by_compound.loc["Compound A", "Peak_Area"], 8525.0)
        self.assertAlmostEqual(
            by_compound.loc["Compound A", "toxpi"],
            normalized.loc[normalized["compound"].eq("Compound A"), "toxpi"].mean(),
        )

    def test_calculate_pbm_toxpi_excludes_missing_pbm_from_final_ranking(self):
        data = pd.DataFrame({
            "compound": ["PBM missing", "Complete"],
            "Peak_Area": [1e9, 1e3],
            "Scores": [float("nan"), 1.0],
            "DF": [1.0, 0.1],
        })
        result = calculate_pbm_toxpi(
            data, PBMToxPiConfig(candidate_top_n=2, display_top_n=2, robustness_enabled=False)
        )
        self.assertEqual(result.final_ranking["compound"].tolist(), ["Complete"])
        self.assertEqual(
            result.excluded_rows.set_index("compound").loc["PBM missing", "exclusion_reason"],
            "PBM score missing",
        )

    def test_calculate_pbm_toxpi_audits_missing_compound_name(self):
        data = pd.DataFrame({
            "compound": ["", "Complete"],
            "Peak_Area": [100.0, 10.0],
            "Scores": [1.0, 1.0],
            "DF": [0.5, 0.5],
        })

        result = calculate_pbm_toxpi(
            data, PBMToxPiConfig(candidate_top_n=2, display_top_n=2, robustness_enabled=False)
        )

        self.assertEqual(result.final_ranking["compound"].tolist(), ["Complete"])
        self.assertEqual(len(result.excluded_rows), 1)
        self.assertEqual(result.excluded_rows.loc[0, "exclusion_reason"], "Compound name missing")

    def test_calculate_pbm_toxpi_excludes_invalid_metrics_and_pov_state(self):
        data = pd.DataFrame({
            "compound": ["No PA", "Bad DF", "Pov failed", "Incomplete", "Complete"],
            "Peak_Area": [float("nan"), 10.0, 10.0, 10.0, 10.0],
            "Scores": [1.0, 1.0, 1.0, 1.0, 1.0],
            "DF": [0.5, 1.2, 0.5, 0.5, 0.5],
            "Pov_LRTP_Status": ["ok", "ok", "error", "ok", "ok"],
            "Pov_LRTP_model_input_complete": [True, True, True, False, True],
        })
        result = calculate_pbm_toxpi(data, PBMToxPiConfig(candidate_top_n=5, display_top_n=5))
        self.assertEqual(result.final_ranking["compound"].tolist(), ["Complete"])
        self.assertEqual(len(result.excluded_rows), 4)

    def test_pbm_toxpi_input_retains_pov_lrtp_eligibility_metadata(self):
        df_table = pd.DataFrame(
            {"compound": ["Compound A"], "Peak_Area": [10.0], "DF": [0.5]}
        )
        pov_lrtp = pd.DataFrame(
            {
                "Name": ["Compound A"],
                "Scores": [1.0],
                "Status": ["error"],
                "model_input_complete": [False],
                "Error": ["Model inputs unavailable"],
            }
        )

        toxpi_input = build_pbm_toxpi_input(df_table, pov_lrtp)

        self.assertEqual(toxpi_input.loc[0, "Pov_LRTP_Status"], "error")
        self.assertFalse(toxpi_input.loc[0, "Pov_LRTP_model_input_complete"])
        self.assertEqual(toxpi_input.loc[0, "Pov_LRTP_Error"], "Model inputs unavailable")

    def test_two_stage_toxpi_selects_global_candidates_then_renormalizes_source_metrics(self):
        toxpi_input = pd.DataFrame(
            {
                "compound": ["A", "B", "C", "D"],
                "Peak_Area": [1e9, 1e8, 1e6, 1e3],
                "Scores": [1.0, 9.0, 8.0, 2.0],
                "DF": [0.2, 0.9, 0.8, 0.1],
            }
        )
        config = PBMToxPiConfig(candidate_top_n=3, display_top_n=2)

        result = calculate_pbm_toxpi(toxpi_input, config=config)

        self.assertIsInstance(result, PBMToxPiResult)
        self.assertEqual(len(result.global_screen), 4)
        self.assertEqual(len(result.candidate_normalized), 3)
        self.assertEqual(len(result.final_ranking), 3)
        self.assertEqual(len(result.display_rows), 2)
        candidates = set(result.candidate_normalized["compound"])
        source = result.source_metrics.set_index("compound")
        self.assertEqual(
            result.candidate_normalized.set_index("compound").loc["B", "Scores"],
            source.loc["B", "Scores"],
        )
        self.assertEqual(candidates, set(result.global_screen.head(3)["compound"]))
        self.assertEqual(
            result.display_rows["compound"].tolist(),
            result.final_ranking.head(2)["compound"].tolist(),
        )
        stage_one = result.global_screen.set_index("compound")
        stage_two = result.candidate_normalized.set_index("compound")
        self.assertNotEqual(
            stage_one.loc["B", "norm_peak_area"],
            stage_two.loc["B", "norm_peak_area"],
        )

    def test_two_stage_toxpi_caps_limits_for_small_inputs(self):
        data = pd.DataFrame(
            {"compound": ["A", "B"], "Peak_Area": [100, 10], "Scores": [2, 1], "DF": [1, 0]}
        )
        result = calculate_pbm_toxpi(
            data,
            config=PBMToxPiConfig(candidate_top_n=100, display_top_n=20),
        )
        self.assertEqual(len(result.final_ranking), 2)
        self.assertEqual(len(result.display_rows), 2)
        self.assertEqual(result.effective_candidate_top_n, 2)
        self.assertEqual(result.effective_display_top_n, 2)

    def test_two_stage_toxpi_normalizes_custom_weights_and_uses_them_in_both_stages(self):
        data = pd.DataFrame(
            {"compound": ["PA", "PBM"], "Peak_Area": [1e8, 1e2], "Scores": [1, 10], "DF": [0.5, 0.5]}
        )
        pa_result = calculate_pbm_toxpi(
            data,
            config=PBMToxPiConfig(candidate_top_n=2, display_top_n=2, weights={"peak_area": 8, "pbm": 1, "df": 1}),
        )
        pbm_result = calculate_pbm_toxpi(
            data,
            config=PBMToxPiConfig(candidate_top_n=2, display_top_n=2, weights={"peak_area": 1, "pbm": 8, "df": 1}),
        )
        self.assertEqual(pa_result.final_ranking.loc[0, "compound"], "PA")
        self.assertEqual(pbm_result.final_ranking.loc[0, "compound"], "PBM")
        self.assertAlmostEqual(sum(pa_result.normalized_weights.values()), 1.0)

    def test_custom_weights_can_change_stage_one_candidate_membership(self):
        data = pd.DataFrame(
            {
                "compound": ["PA", "PBM", "Mid"],
                "Peak_Area": [1e9, 1e2, 1e5],
                "Scores": [1.0, 10.0, 5.0],
                "DF": [0.5, 0.5, 0.5],
            }
        )

        pa_result = calculate_pbm_toxpi(
            data,
            PBMToxPiConfig(
                candidate_top_n=1,
                display_top_n=1,
                weights={"peak_area": 8, "pbm": 1, "df": 1},
                robustness_enabled=False,
            ),
        )
        pbm_result = calculate_pbm_toxpi(
            data,
            PBMToxPiConfig(
                candidate_top_n=1,
                display_top_n=1,
                weights={"peak_area": 1, "pbm": 8, "df": 1},
                robustness_enabled=False,
            ),
        )

        self.assertEqual(pa_result.candidate_normalized["compound"].tolist(), ["PA"])
        self.assertEqual(pbm_result.candidate_normalized["compound"].tolist(), ["PBM"])

    def test_two_stage_toxpi_rejects_invalid_settings(self):
        with self.assertRaisesRegex(ValueError, "positive"):
            PBMToxPiConfig(weights={"peak_area": 0, "pbm": 0, "df": 0})
        with self.assertRaisesRegex(ValueError, "cannot exceed"):
            PBMToxPiConfig(candidate_top_n=10, display_top_n=20)

    def test_toxpi_evidence_plot_limits_are_validated_and_exported(self):
        config = PBMToxPiConfig(
            evidence_per_compound_top_n=7,
            evidence_global_use_top_n=23,
        )
        result = calculate_pbm_toxpi(
            pd.DataFrame(
                {
                    "compound": ["A"],
                    "Peak_Area": [100.0],
                    "Scores": [2.0],
                    "DF": [1.0],
                }
            ),
            config=config,
        )
        settings = result.settings_table().set_index("setting")["value"]

        self.assertEqual(settings["evidence_per_compound_top_n"], 7)
        self.assertEqual(settings["evidence_global_use_top_n"], 23)
        for field in ("evidence_per_compound_top_n", "evidence_global_use_top_n"):
            with self.subTest(field=field), self.assertRaisesRegex(ValueError, "at least 1"):
                PBMToxPiConfig(**{field: 0})

    def test_two_stage_toxpi_rejects_empty_weights(self):
        with self.assertRaisesRegex(ValueError, "positive"):
            PBMToxPiConfig(weights={})

    def test_two_stage_toxpi_rejects_non_finite_weights(self):
        for invalid in (float("nan"), float("inf"), float("-inf")):
            with self.subTest(weight=invalid), self.assertRaisesRegex(ValueError, "finite"):
                PBMToxPiConfig(
                    weights={"peak_area": invalid, "pbm": 0.4, "df": 0.2}
                )

    def test_two_stage_toxpi_rejects_non_finite_perturbation(self):
        for invalid in (float("nan"), float("inf"), float("-inf")):
            with self.subTest(perturbation=invalid), self.assertRaisesRegex(
                ValueError,
                "finite",
            ):
                PBMToxPiConfig(perturbation_fraction=invalid)

    def test_two_stage_toxpi_breaks_score_ties_deterministically(self):
        scored = pd.DataFrame(
            {
                "compound": ["A", "B", "C"],
                "Peak_Area": [10, 100, 100],
                "toxpi": [0.5, 0.5, 0.5],
            }
        )

        result = workflow._sort_toxpi_stage(scored, "toxpi", "final_rank")

        self.assertEqual(result["compound"].tolist(), ["B", "C", "A"])
        self.assertEqual(result["final_rank"].tolist(), [1, 2, 3])

    def test_screening_workbook_contains_two_stage_and_robustness_sheets(self):
        expected = {
            "ToxPi_Global_Screen",
            "ToxPi_Normalized",
            "ToxPi_Results",
            "ToxPi_Display",
            "ToxPi_Settings",
            "ToxPi_Robustness",
            "ToxPi_Robust_Stats",
        }

        self.assertTrue(expected.issubset(EXPECTED_WORKBOOK_SHEETS))

    def test_screening_workbook_and_page_expose_toxpi_exclusions(self):
        self.assertIn("ToxPi_Excluded", EXPECTED_WORKBOOK_SHEETS)
        page_text = Path("pages/0_综合筛查流程.py").read_text(encoding="utf-8")
        self.assertIn('"toxpi_excluded": toxpi_result.excluded_rows', page_text)
        self.assertIn('st.subheader("ToxPi_Excluded")', page_text)

    def test_comprehensive_page_handles_legacy_downstream_state_without_toxpi_exclusions(self):
        page_text = Path("pages/0_综合筛查流程.py").read_text(encoding="utf-8")
        self.assertIn(
            'show_dataframe(downstream_state.get("toxpi_excluded", pd.DataFrame()))',
            page_text,
        )

    def test_comprehensive_page_exposes_shared_axis_toxpi_and_robustness_controls(self):
        page_text = Path("pages/0_综合筛查流程.py").read_text(encoding="utf-8")
        for token in (
            "ScreeningAxisRanges(",
            "PBMToxPiConfig(",
            "candidate_top_n",
            "display_top_n",
            "perturbation_fraction",
            "robustness_enabled",
            "ToxPi_Global_Screen",
            "ToxPi_Robustness",
            "cp_screening_settings_signature",
            "cp_screening_robustness_png",
            "cp_screening_robustness_pdf",
            "hashlib.sha256",
            "normalized_weights",
            "effective_display_top_n",
        ):
            self.assertIn(token, page_text)

    def test_comprehensive_page_clears_downstream_plot_artifacts_before_regeneration(self):
        page_text = Path("pages/0_综合筛查流程.py").read_text(encoding="utf-8")
        state_keys_token = "DOWNSTREAM_PLOT_STATE_KEYS = ("
        self.assertIn(state_keys_token, page_text)
        state_keys_start = page_text.index(state_keys_token)
        state_keys_end = page_text.index(")", state_keys_start)
        state_keys_text = page_text[state_keys_start:state_keys_end]
        for key in (
            "cp_screening_bar_png",
            "cp_screening_bar_pdf",
            "cp_screening_radial_png",
            "cp_screening_radial_pdf",
            "cp_screening_radial_plot_version",
            "cp_screening_robustness_png",
            "cp_screening_robustness_pdf",
        ):
            self.assertIn(key, state_keys_text)

        result_state_index = page_text.index('st.session_state["cp_screening_downstream"] = {')
        clear_index = page_text.index(
            "for key in DOWNSTREAM_PLOT_STATE_KEYS:",
            result_state_index,
        )
        display_plot_index = page_text.index(
            "if not toxpi_result.display_rows.empty",
            result_state_index,
        )
        robustness_plot_index = page_text.index(
            "if not toxpi_result.robustness_correlations.empty",
            result_state_index,
        )
        self.assertLess(result_state_index, clear_index)
        self.assertLess(clear_index, display_plot_index)
        self.assertLess(clear_index, robustness_plot_index)
        clear_block = page_text[clear_index:display_plot_index]
        self.assertIn("st.session_state.pop(key, None)", clear_block)

    def test_toxpi_robustness_is_reproducible_and_uses_configured_display_top_n(self):
        data = pd.DataFrame(
            {
                "compound": [f"C{i}" for i in range(6)],
                "Peak_Area": [1e8, 1e7, 1e6, 1e5, 1e4, 1e3],
                "Scores": [1, 4, 3, 6, 2, 5],
                "DF": [0.9, 0.2, 0.8, 0.4, 0.7, 0.1],
            }
        )
        config = PBMToxPiConfig(
            candidate_top_n=6,
            display_top_n=2,
            perturbation_fraction=0.35,
            n_iter=40,
            seed=77,
        )

        first = run_pbm_toxpi_robustness(calculate_pbm_toxpi(data, config), config)
        second = run_pbm_toxpi_robustness(calculate_pbm_toxpi(data, config), config)

        pd.testing.assert_frame_equal(first.robustness_summary, second.robustness_summary)
        pd.testing.assert_frame_equal(first.robustness_correlations, second.robustness_correlations)
        self.assertEqual(first.robustness_stats.loc[0, "perturbation_fraction"], 0.35)
        self.assertEqual(first.robustness_stats.loc[0, "display_top_n"], 2)
        self.assertTrue(first.robustness_summary["top_n_frequency_percent"].between(0, 100).all())
        self.assertAlmostEqual(first.robustness_summary["top_n_frequency_percent"].sum(), 200.0)

    def test_toxpi_robustness_handles_missing_indicators(self):
        data = pd.DataFrame(
            {
                "compound": ["A", "B", "C", "D"],
                "Peak_Area": [1e8, float("nan"), 1e5, 1e3],
                "Scores": [1.0, 8.0, float("nan"), 4.0],
                "DF": [0.9, 0.7, float("nan"), 0.2],
            }
        )
        config = PBMToxPiConfig(
            candidate_top_n=4,
            display_top_n=2,
            n_iter=40,
            seed=11,
        )

        result = calculate_pbm_toxpi(data, config)

        self.assertEqual(len(result.robustness_correlations), 40)
        self.assertTrue(result.robustness_correlations["spearman_rho"].notna().all())
        self.assertAlmostEqual(result.robustness_summary["top_n_frequency_percent"].sum(), 200.0)

    def test_toxpi_robustness_plot_renders_rank_correlation_distribution(self):
        data = pd.DataFrame(
            {
                "compound": ["A", "B", "C"],
                "Peak_Area": [1e6, 1e4, 1e2],
                "Scores": [1, 5, 3],
                "DF": [0.9, 0.2, 0.6],
            }
        )
        config = PBMToxPiConfig(candidate_top_n=3, display_top_n=2, n_iter=10, seed=7)
        result = run_pbm_toxpi_robustness(calculate_pbm_toxpi(data, config), config)

        fig = generate_pbm_toxpi_robustness_plot(result)

        try:
            self.assertEqual(fig.axes[0].get_title(), "ToxPi Rank Robustness")
            self.assertEqual(
                fig.axes[0].get_xlabel(),
                "Spearman correlation with baseline ranking",
            )
        finally:
            plt.close(fig)

    def test_toxpi_robustness_plot_rejects_empty_correlations(self):
        data = pd.DataFrame(
            {
                "compound": ["A", "B"],
                "Peak_Area": [100, 10],
                "Scores": [2, 1],
                "DF": [1, 0],
            }
        )
        result = calculate_pbm_toxpi(
            data,
            PBMToxPiConfig(
                candidate_top_n=2,
                display_top_n=2,
                robustness_enabled=False,
            ),
        )

        with self.assertRaisesRegex(ValueError, "Robustness correlations are empty"):
            generate_pbm_toxpi_robustness_plot(result)

    def test_figure_pair_serializer_closes_figure_when_serialization_raises(self):
        self.assertTrue(hasattr(workflow, "figure_to_png_pdf_bytes"))
        serialize = workflow.figure_to_png_pdf_bytes
        original_png = workflow.figure_to_png_bytes
        original_pdf = workflow.figure_to_pdf_bytes

        for failing_format in ("png", "pdf"):
            with self.subTest(failing_format=failing_format):
                fig = plt.figure()

                def succeed(_fig):
                    return io.BytesIO(b"ok")

                def fail(_fig):
                    raise RuntimeError(f"{failing_format} serialization failed")

                workflow.figure_to_png_bytes = fail if failing_format == "png" else succeed
                workflow.figure_to_pdf_bytes = fail if failing_format == "pdf" else succeed
                try:
                    with self.assertRaisesRegex(RuntimeError, "serialization failed"):
                        serialize(fig)
                    self.assertFalse(plt.fignum_exists(fig.number))
                finally:
                    workflow.figure_to_png_bytes = original_png
                    workflow.figure_to_pdf_bytes = original_pdf
                    plt.close(fig)

    def test_comprehensive_page_uses_closing_serializer_for_all_toxpi_figures(self):
        page_text = Path("pages/0_综合筛查流程.py").read_text(encoding="utf-8")

        self.assertEqual(page_text.count("figure_to_png_pdf_bytes("), 3)
        plot_block = page_text.split("def refresh_toxpi_radial_plot", 1)[1].split(
            'st.session_state["cp_screening_workbook"]', 1
        )[0]
        self.assertNotIn("figure_to_png_bytes(", plot_block)
        self.assertNotIn("figure_to_pdf_bytes(", plot_block)

    def test_calculate_pbm_toxpi_runs_enabled_robustness_analysis(self):
        data = pd.DataFrame(
            {
                "compound": ["A", "B", "C"],
                "Peak_Area": [1000, 100, 10],
                "Scores": [1, 3, 2],
                "DF": [0.8, 0.2, 0.5],
            }
        )

        result = calculate_pbm_toxpi(
            data,
            PBMToxPiConfig(candidate_top_n=3, display_top_n=2, n_iter=12),
        )

        self.assertEqual(len(result.robustness_correlations), 12)
        self.assertEqual(len(result.robustness_summary), 3)
        self.assertEqual(len(result.robustness_stats), 1)

    def test_calculate_pbm_toxpi_leaves_robustness_tables_empty_when_disabled(self):
        data = pd.DataFrame(
            {
                "compound": ["A", "B"],
                "Peak_Area": [100, 10],
                "Scores": [2, 1],
                "DF": [1, 0],
            }
        )

        result = calculate_pbm_toxpi(
            data,
            PBMToxPiConfig(
                candidate_top_n=2,
                display_top_n=2,
                robustness_enabled=False,
            ),
        )

        self.assertTrue(result.robustness_summary.empty)
        self.assertTrue(result.robustness_stats.empty)
        self.assertTrue(result.robustness_correlations.empty)

    def test_calculate_pbm_toxpi_keeps_one_candidate_without_running_robustness(self):
        data = pd.DataFrame(
            {
                "compound": ["Only"],
                "Peak_Area": [100],
                "Scores": [2],
                "DF": [1],
            }
        )

        result = calculate_pbm_toxpi(
            data,
            PBMToxPiConfig(candidate_top_n=1, display_top_n=1),
        )

        self.assertEqual(result.final_ranking["compound"].tolist(), ["Only"])
        self.assertTrue(result.robustness_summary.empty)
        self.assertTrue(result.robustness_stats.empty)
        self.assertTrue(result.robustness_correlations.empty)

    def test_pbm_scores_are_normalized_in_positive_direction_for_toxpi(self):
        toxpi_input = pd.DataFrame(
            {
                "compound": ["High", "Middle", "Low"],
                "Peak_Area": [300.0, 200.0, 100.0],
                "Scores": [9.0, 5.0, 1.0],
                "DF": [1.0, 0.5, 0.0],
            }
        )

        result = calculate_pbm_toxpi(toxpi_input)
        normalized = result.candidate_normalized
        toxpi_results = result.final_ranking

        normalized_by_compound = normalized.set_index("compound")
        self.assertGreater(
            normalized_by_compound.loc["High", "norm_pbm"],
            normalized_by_compound.loc["Low", "norm_pbm"],
        )
        self.assertEqual(toxpi_results["compound"].tolist(), ["High", "Middle", "Low"])
        self.assertGreater(toxpi_results.loc[0, "toxpi"], toxpi_results.loc[1, "toxpi"])
        self.assertGreater(toxpi_results.loc[1, "toxpi"], toxpi_results.loc[2, "toxpi"])

    def test_toxpi_radial_plot_rows_are_limited_for_web_preview(self):
        toxpi_results = pd.DataFrame(
            {
                "compound": [f"Compound {index:03d}" for index in range(30)],
                "toxpi": np.linspace(1.0, 0.1, 30),
                "norm_peak_area": np.linspace(1.0, 0.1, 30),
                "norm_pbm": np.linspace(0.9, 0.0, 30),
                "norm_df": np.linspace(0.8, 0.0, 30),
            }
        )
        toxpi_results.attrs["toxic_cols"] = ["peak_area", "pbm", "df"]

        plot_rows, omitted_count = limit_toxpi_plot_rows(toxpi_results, max_compounds=15)

        self.assertEqual(len(plot_rows), 15)
        self.assertEqual(omitted_count, 15)
        self.assertEqual(plot_rows["compound"].iloc[0], "Compound 000")
        self.assertEqual(plot_rows["compound"].iloc[-1], "Compound 014")
        self.assertEqual(plot_rows.attrs["toxic_cols"], ["peak_area", "pbm", "df"])

    def test_screening_workbook_contains_expected_sheets_even_for_empty_tables(self):
        self.assertIn("Group_Area_Raw_Long", EXPECTED_WORKBOOK_SHEETS)
        self.assertIn("Group_Area_Mean_By_Sample", EXPECTED_WORKBOOK_SHEETS)
        workbook_buffer = build_screening_workbook(
            {
                "Input_Check": pd.DataFrame({"status": ["ok"]}),
                "ToxPi_Results": pd.DataFrame({"compound": ["A"], "toxpi": [0.8]}),
            }
        )

        workbook_buffer.seek(0)
        workbook = load_workbook(io.BytesIO(workbook_buffer.getvalue()), read_only=True)
        self.assertEqual(workbook.sheetnames, EXPECTED_WORKBOOK_SHEETS)

    def test_warning_stage_helper_preserves_existing_stage_column(self):
        helper = getattr(workflow, "with_warning_stage", None)
        self.assertIsNotNone(helper, "with_warning_stage should be implemented")

        existing_stage = pd.DataFrame(
            {
                "message": ["already staged", "needs fallback"],
                "stage": ["PubChem", ""],
            }
        )
        staged_existing = helper(existing_stage, "identifier_warnings")
        self.assertEqual(staged_existing.columns.tolist(), ["stage", "message"])
        self.assertEqual(staged_existing["stage"].tolist(), ["PubChem", "identifier_warnings"])

        missing_stage = pd.DataFrame({"message": ["plain warning"]})
        staged_missing = helper(missing_stage, "epi_errors")
        self.assertEqual(staged_missing.columns.tolist(), ["stage", "message"])
        self.assertEqual(staged_missing.loc[0, "stage"], "epi_errors")

    def test_fourth_page_no_longer_exposes_epa_echa_combined_plot(self):
        page_text = Path("pages/4_化合物用途查询.py").read_text(encoding="utf-8")

        self.assertNotIn("EPA_ECHA_Combined", page_text)
        self.assertNotIn("generate_combined_use_rose_plot", page_text)
        self.assertNotIn("build_epa_echa_combined_rose_data", page_text)

    def test_fourth_page_uses_single_puc_distribution_instead_of_rose_plot(self):
        page_text = Path("pages/4_化合物用途查询.py").read_text(encoding="utf-8")

        self.assertIn("extract_top_product_use_category_data", page_text)
        self.assertIn("EPA CompTox Product-Use Category Distribution", page_text)
        self.assertIn("EPA_Product_Use_Category_Distribution", page_text)
        self.assertIn("PRODUCT_USE_CATEGORY_OTHERS_NOTE", page_text)
        self.assertNotIn("EPA CompTox Product-Use Category Rose Plot", page_text)
        self.assertNotIn("EPA_Product_Use_Category_Rose_Plot", page_text)

    def test_comprehensive_screening_page_front_half_figures_are_rendered(self):
        page_path = Path("pages/0_综合筛查流程.py")
        page_text = page_path.read_text(encoding="utf-8")

        self.assertIn("render_front_half_figures(front_state)", page_text)
        self.assertNotIn("point_screening_results", page_text)
        self.assertNotIn("POINT_FRONT_HALF_FIGURES", page_text)
        self.assertIn("Group_Area_Mean", page_text)
        self.assertIn("build_group_area_mean_by_sample", page_text)
        self.assertIn("def with_warning_stage(", page_text)
        self.assertIn("with_warning_stage(table, key)", page_text)
        self.assertNotIn("    with_warning_stage,\n", page_text)
        workflow_tables_text = page_text.split("def workflow_tables", 1)[1]
        self.assertNotIn('insert(0, "stage"', workflow_tables_text)
        self.assertIn("st.image", page_text)
        self.assertIn("category_percent_donut_with_total", page_text)
        self.assertIn("compound_bubble_plot", page_text)
        self.assertIn("VanKrevelen", page_text)
        self.assertIn("PER_SAMPLE_FRONT_HALF_FIGURES", page_text)
        self.assertIn("SUMMARY_FRONT_HALF_FIGURES", page_text)
        self.assertIn("summary_figure_paths", page_text)
        self.assertIn("save_boxplot_log_transformed", page_text)
        per_sample_figures = page_text.split("PER_SAMPLE_FRONT_HALF_FIGURES", 1)[1].split("]", 1)[0]
        self.assertNotIn("boxplot_log_transformed", per_sample_figures)
        self.assertNotIn("TOXPI_RADIAL_MAX_COMPOUNDS", page_text)
        self.assertIn("TOXPI_RADIAL_PLOT_VERSION", page_text)
        self.assertNotIn("limit_toxpi_plot_rows", page_text)
        self.assertIn("generate_r_style_toxpi_plot", page_text)
        self.assertIn("import importlib", page_text)
        self.assertIn("import src.toxpi_calc as toxpi_calc", page_text)
        self.assertIn("importlib.reload(toxpi_calc)", page_text)
        self.assertNotIn("from src.toxpi_calc import generate_r_style_toxpi_plot", page_text)
        self.assertIn("label_wrap_width=20", page_text)
        self.assertIn("refresh_toxpi_radial_plot", page_text)
        self.assertIn('"cp_screening_radial_plot_version"', page_text)
        self.assertNotIn("generate_multi_toxpi_plot", page_text)
        self.assertIn("def render_sample_mapping_tabs(samples):", page_text)
        self.assertIn("mapping_tabs = st.tabs", page_text)
        self.assertIn("normalize_samples_for_mappings", page_text)
        self.assertIn("sample_mappings", page_text)
        self.assertIn('front_state["representative_table"]', page_text)


if __name__ == "__main__":
    unittest.main()
