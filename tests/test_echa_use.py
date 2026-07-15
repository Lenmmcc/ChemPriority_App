import io
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

from src import echa_use
from src.query_cache import use_cache_path


def _resolution():
    return {
        "echa_id": "100.001.133",
        "matched_name": "Example",
        "matched_cas": "50-00-0",
        "matched_ec": "200-001-8",
        "status": "matched",
        "message": "",
    }


def _dossier():
    return {
        "asset_external_id": "DISS-1",
        "registration_number": "01-0000000000-00",
        "registration_status": "Active",
        "registration_date": "",
        "last_updated_date": "",
        "dossier_subtype": "",
        "registration_role": "",
    }


def _candidate(raw_use="Industrial use", use_cn="工业用途"):
    return {
        "use_cn": use_cn,
        "use_en": raw_use,
        "raw_use": raw_use,
        "echa_use_section": "Uses at industrial sites",
        "use_phase": "industrial",
        "evidence_count": 1,
        "source_type": "reach_use",
        "source": "ECHA REACH dossier",
        "dossier_asset_id": "DISS-1",
        "registration_number": "01-0000000000-00",
        "registration_status": "Active",
        "dossier_subtype": "",
        "registration_role": "",
        "last_updated_date": "",
        "record_url": "https://example.test/record",
        "dossier_url": "https://example.test/dossier",
    }


class EchaUseSummaryTests(unittest.TestCase):
    @patch("src.echa_use.extract_dossier_use_candidates")
    @patch("src.echa_use.fetch_dossier_html", return_value="<html></html>")
    @patch("src.echa_use.fetch_reach_dossiers", return_value=([_dossier()], []))
    @patch("src.echa_use.resolve_substance")
    def test_batch_keeps_divergent_name_and_smiles_echa_groups_with_conflict_warning(
        self, resolve_substance, fetch_dossiers, fetch_html, extract_candidates
    ):
        resolve_substance.side_effect = [
            {**_resolution(), "echa_id": "100.001.999", "status": "使用输入 ECHA ID"},
            {**_resolution(), "echa_id": "100.001.111", "status": "通过 compound 匹配"},
            {**_resolution(), "echa_id": "100.001.222", "status": "通过 smiles 匹配"},
        ]
        extract_candidates.side_effect = [
            [_candidate("Input use", "输入标识用途")],
            [_candidate("Name use", "名称用途")],
            [_candidate("SMILES use", "结构用途")],
        ]

        summary_df, candidates_df, dossiers_df, warnings_df = echa_use.run_echa_use_batch(
            pd.DataFrame(
                [
                    {
                        "compound": "Example name",
                        "cas": "50-00-0",
                        "ec": "200-001-8",
                        "smiles": "CCO",
                        "echa_id": "100.001.999",
                    }
                ]
            ),
            delay_seconds=0,
        )

        self.assertEqual(resolve_substance.call_count, 3)
        input_query = resolve_substance.call_args_list[0].args[0]
        name_query = resolve_substance.call_args_list[1].args[0]
        smiles_query = resolve_substance.call_args_list[2].args[0]
        self.assertEqual(input_query["echa_id"], "100.001.999")
        self.assertEqual(name_query["compound"], "Example name")
        self.assertEqual(name_query["smiles"], "")
        self.assertEqual(name_query["cas"], "")
        self.assertEqual(name_query["ec"], "")
        self.assertEqual(name_query["echa_id"], "")
        self.assertEqual(smiles_query["compound"], "")
        self.assertEqual(smiles_query["smiles"], "CCO")
        self.assertEqual(smiles_query["cas"], "")
        self.assertEqual(smiles_query["ec"], "")
        self.assertEqual(smiles_query["echa_id"], "")
        self.assertEqual(summary_df["query_source"].tolist(), ["输入标识", "名称", "SMILES"])
        self.assertEqual(candidates_df["echa_id"].tolist(), ["100.001.999", "100.001.111", "100.001.222"])
        self.assertEqual(candidates_df["query_source"].tolist(), ["输入标识", "名称", "SMILES"])
        self.assertEqual(candidates_df["is_primary_identity"].tolist(), [True, False, False])
        self.assertEqual(dossiers_df["query_source"].tolist(), ["输入标识", "名称", "SMILES"])
        self.assertEqual(dossiers_df["query_value"].tolist(), ["100.001.999", "Example name", "CCO"])
        self.assertEqual(dossiers_df["is_primary_identity"].tolist(), [True, False, False])
        conflict_warnings = warnings_df[warnings_df["stage"].eq("identity_conflict")]
        self.assertEqual(len(conflict_warnings), 1)
        self.assertEqual(conflict_warnings.iloc[0]["query_source"], "名称 | SMILES")

    @patch("src.echa_use.urllib.request.urlopen")
    def test_get_text_uses_query_cache(self, urlopen):
        response = unittest.mock.MagicMock()
        response.read.return_value = b'{"items":[]}'
        urlopen.return_value.__enter__.return_value = response

        with tempfile.TemporaryDirectory() as tmpdir:
            with use_cache_path(Path(tmpdir) / "queries.sqlite3"):
                first = echa_use._get_text(
                    "api-substance/v1/substance",
                    params={"searchText": "ethanol"},
                    base_url="https://example.test/",
                    timeout=1,
                )
                second = echa_use._get_text(
                    "api-substance/v1/substance",
                    params={"searchText": "ethanol"},
                    base_url="https://example.test/",
                    timeout=1,
                )

        self.assertEqual(first, '{"items":[]}')
        self.assertEqual(second, '{"items":[]}')
        urlopen.assert_called_once()

    @patch("src.echa_use.extract_dossier_use_candidates")
    @patch("src.echa_use.fetch_dossier_html", return_value="<html></html>")
    @patch("src.echa_use.fetch_reach_dossiers")
    @patch("src.echa_use.resolve_substance", return_value=_resolution())
    def test_summary_omits_top_ranking_columns(self, resolve, fetch_dossiers, fetch_html, extract_candidates):
        fetch_dossiers.return_value = ([_dossier()], [])
        extract_candidates.return_value = [_candidate(), _candidate("Consumer use", "消费者用途")]

        input_df = pd.DataFrame([{"compound": "Example", "cas": "50-00-0"}])
        summary_df, candidates_df, dossiers_df, errors_df = echa_use.run_echa_use_batch(
            input_df,
            delay_seconds=0,
        )

        self.assertTrue(errors_df.empty)
        self.assertEqual(len(candidates_df), 2)
        self.assertEqual(summary_df.loc[0, "query_status"], "查询完成")
        self.assertEqual(summary_df.loc[0, "ECHA_dossier数量"], 1)
        for column in ("前五用途", "用途来源", "用途1", "用途1_英文证据", "用途1_证据数量"):
            self.assertNotIn(column, summary_df.columns)

        workbook = echa_use.build_result_workbook(
            input_df,
            summary_df=summary_df,
            candidates_df=candidates_df,
            dossiers_df=dossiers_df,
            errors_df=errors_df,
        )
        book = load_workbook(io.BytesIO(workbook.getvalue()), read_only=True)
        self.assertIn("ECHA_Use_Summary", book.sheetnames)
        self.assertNotIn("ECHA_Top5_Use_Summary", book.sheetnames)
        self.assertIn("ECHA_Uses_Reported", book.sheetnames)
        self.assertIn("ECHA_Reported_Pie_Data", book.sheetnames)
        self.assertNotIn("ECHA_All_Use_Candidates", book.sheetnames)
        echa_pie = pd.read_excel(
            io.BytesIO(workbook.getvalue()), sheet_name="ECHA_Reported_Pie_Data"
        )
        self.assertEqual(len(echa_pie), len(input_df.drop_duplicates("compound")))

    def test_empty_summary_template_has_no_top_ranking_columns(self):
        template = echa_use.build_empty_summary_template(pd.DataFrame([{"compound": "Example"}]))

        for column in ("前五用途", "用途来源", "用途1", "用途1_英文证据", "用途1_证据数量"):
            self.assertNotIn(column, template.columns)


if __name__ == "__main__":
    unittest.main()
