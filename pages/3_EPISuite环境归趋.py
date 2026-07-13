import hashlib
import io
import os
import sys

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(CURRENT_DIR)
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from src.episuite_io import (  # noqa: E402
    DEFAULT_EPI_WEB_API,
    FATE_ENDPOINTS,
    REQUIRED_COLUMNS,
    build_epi_web_result_tables,
    build_input_zip,
    build_result_workbook,
    input_columns_for_display,
    make_template_file,
    merge_results_with_input,
    normalize_input_columns,
    parse_uploaded_result,
    run_epi_web_batch,
    slim_epi_report_columns,
    validate_input,
)
from src.query_cache import clear_query_cache, current_cache_path  # noqa: E402
from src.mol_structure_parser import (  # noqa: E402
    prepare_structure_dataframe,
    summarize_structure_preparation,
)


INPUT_CACHE_KEYS = (
    "epi_input_bytes",
    "epi_input_name",
    "epi_input_signature",
)
RESULT_CACHE_KEYS = (
    "epi_web_results",
    "epi_web_errors",
    "epi_raw_results",
    "epi_web_tables",
    "epi_merged_results",
    "epi_parsed_results",
    "epi_parse_warnings",
)

DETAIL_RESULT_SHEETS = [
    ("Properties", "理化性质"),
    ("Degradation", "降解"),
    ("Fate_Transport", "归趋/迁移"),
    ("Bioaccumulation", "富集"),
    ("ECOSAR_Aquatic_Toxicity", "ECOSAR"),
    ("Model_Metadata", "模型元数据"),
]


def clear_result_cache():
    for key in RESULT_CACHE_KEYS:
        st.session_state.pop(key, None)


def clear_cached_input():
    for key in INPUT_CACHE_KEYS:
        st.session_state.pop(key, None)
    clear_result_cache()


def append_structure_preparation_sheet(workbook_buffer, prepared_df):
    workbook_buffer.seek(0)
    workbook = load_workbook(workbook_buffer)
    if "Structure_Preparation" in workbook.sheetnames:
        del workbook["Structure_Preparation"]
    worksheet = workbook.create_sheet("Structure_Preparation")
    export_df = prepared_df.astype(object).where(prepared_df.notna(), None)
    for row in dataframe_to_rows(export_df, index=False, header=True):
        worksheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def render_epi_web_tables(epi_tables):
    if not epi_tables:
        return

    st.subheader("分类结构化结果")
    detail_tabs = st.tabs([label for _, label in DETAIL_RESULT_SHEETS])
    for (sheet_name, _), tab in zip(DETAIL_RESULT_SHEETS, detail_tabs):
        table = epi_tables.get(sheet_name)
        with tab:
            if table is not None and not table.empty:
                st.dataframe(table, use_container_width=True)
            else:
                st.info("该类别没有可展示的结构化结果。")

    raw_table = epi_tables.get("Raw_API_JSON")
    if raw_table is not None and not raw_table.empty:
        with st.expander("审计数据：Raw API JSON", expanded=False):
            preview_cols = [col for col in ["compound", "smiles", "cas", "epi_cas", "epi_smiles", "raw_json"] if col in raw_table.columns]
            st.dataframe(raw_table[preview_cols], use_container_width=True)


def render_structure_preparation_summary(prepared_df):
    summary = summarize_structure_preparation(prepared_df)
    st.caption("结构准备（MOL / SMILES）")
    labels = ["MOL 行", "解析成功", "修复 M END", "SMILES 冲突", "解析失败"]
    values = [
        summary["mol_rows"],
        summary["parsed_success"],
        summary["repaired_m_end"],
        summary["smiles_conflicts"],
        summary["parse_failures"],
    ]
    for column, label, value in zip(st.columns(5), labels, values):
        column.metric(label, value)
    if summary["smiles_conflicts"] or summary["parse_failures"]:
        with st.expander("查看结构准备审计记录", expanded=False):
            mask = prepared_df["smiles_source"].eq("原始 SMILES（与 MOL 冲突）") | prepared_df["parse_status"].eq("解析失败")
            st.dataframe(prepared_df.loc[mask], use_container_width=True)


st.set_page_config(
    page_title="EPI Suite 环境归趋预测 - ChemPriority",
    page_icon="🌊",
    layout="wide",
)


st.title("🌊 EPI Suite 环境归趋预测")
st.caption("上传 compound + smiles 表格，可选提供 cas；通过 EPI Web Suite 网页端 API 批量计算环境归趋指标。")
st.markdown("---")

left_col, right_col = st.columns([2, 1])

with left_col:
    st.subheader("1. 上传 EPI Suite 输入表")
    uploaded_file = st.file_uploader(
        "上传 Excel 文件",
        type=["xlsx", "xls"],
        help="文件至少需要包含 compound 和 smiles 两列；cas 可选。有 cas 时会与 smiles 一起提交，并同时保留估算值与实验/库值。",
    )

with right_col:
    st.subheader("输入模板")
    st.download_button(
        label="下载 Excel 模板",
        data=make_template_file(),
        file_name="EPISuite_Input_Template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

st.subheader("目标环境归趋指标")
st.dataframe(pd.DataFrame(FATE_ENDPOINTS), use_container_width=True)

if uploaded_file is not None:
    uploaded_bytes = uploaded_file.getvalue()
    input_signature = hashlib.sha256(uploaded_bytes).hexdigest()
    if st.session_state.get("epi_input_signature") != input_signature:
        clear_result_cache()
    st.session_state["epi_input_bytes"] = uploaded_bytes
    st.session_state["epi_input_name"] = uploaded_file.name
    st.session_state["epi_input_signature"] = input_signature

cached_input_bytes = st.session_state.get("epi_input_bytes")
cached_input_name = st.session_state.get("epi_input_name")

if cached_input_bytes is None:
    st.info("请先上传包含 compound 和 smiles 的 Excel 文件；cas 可选。")
    st.stop()

st.success(f"已加载输入文件：{cached_input_name}")
if st.button("清空当前数据", key="epi_clear_cached_input"):
    clear_cached_input()
    st.rerun()

try:
    raw_input_df = pd.read_excel(io.BytesIO(cached_input_bytes))
    prepared_input_df = prepare_structure_dataframe(raw_input_df)
except Exception as exc:
    st.error(f"Excel 读取失败：{exc}")
    st.stop()

render_structure_preparation_summary(prepared_input_df)

try:
    input_df = normalize_input_columns(prepared_input_df)
except Exception as exc:
    st.error(f"Excel 读取失败：{exc}")
    st.stop()

is_valid, message = validate_input(input_df)
if not is_valid:
    st.error(message)
    st.dataframe(input_df, use_container_width=True)
    st.stop()

st.success(message)

tab_input, tab_predict, tab_fallback, tab_parse, tab_output = st.tabs(
    ["输入数据", "网页端预测", "备用输入包", "解析外部结果", "结果下载"]
)

with tab_input:
    st.subheader("待预测化合物")
    st.dataframe(input_df[input_columns_for_display(input_df)], use_container_width=True)
    st.metric("化合物数量", len(input_df))

with tab_predict:
    st.subheader("EPI Web Suite 自动预测")
    st.write(
        "点击后，系统会逐个调用 EPI Web Suite 网页端 API，并把结果整理成表格。"
        "若输入包含 cas，会同时提交 smiles 与 cas；结果表会同时显示 selected、estimated 和 experimental 值。"
    )

    api_url = st.text_input(
        "EPI Web Suite API 地址",
        value=DEFAULT_EPI_WEB_API,
        help="默认使用 EPA/EPI Web Suite 当前网页端 API。部署时服务器需要能访问该地址。",
    )
    col_timeout, col_delay = st.columns(2)
    with col_timeout:
        timeout_seconds = st.number_input("单个化合物超时时间（秒）", min_value=10, max_value=300, value=90, step=10)
    with col_delay:
        delay_seconds = st.number_input("请求间隔（秒）", min_value=0.0, max_value=5.0, value=0.2, step=0.1)

    with st.expander("加速设置", expanded=False):
        cache_enabled = st.checkbox("启用本地查询缓存", value=True, key="epi_query_cache_enabled")
        max_workers = st.number_input(
            "并发数",
            min_value=1,
            max_value=8,
            value=3,
            step=1,
            key="epi_query_max_workers",
            help="默认 3 个并发请求；遇到外部服务限流时可调低。",
        )
        st.caption(f"缓存文件：{current_cache_path()}")
        if st.button("清理本地查询缓存", key="epi_clear_query_cache"):
            clear_query_cache()
            st.success("本地查询缓存已清理。")

    if st.button("开始网页端预测", type="primary"):
        progress_bar = st.progress(0)
        status_box = st.empty()

        def update_progress(done, total, compound):
            progress_bar.progress(done / total)
            status_box.info(f"正在处理：{compound} ({done}/{total})")

        with st.spinner("正在调用 EPI Web Suite，请等待..."):
            web_results, raw_results, web_errors = run_epi_web_batch(
                input_df,
                api_url=api_url,
                timeout=int(timeout_seconds),
                delay_seconds=float(delay_seconds),
                max_workers=int(max_workers),
                cache_enabled=bool(cache_enabled),
                progress_callback=update_progress,
            )
            web_tables = build_epi_web_result_tables(web_results, raw_results, web_errors)

        st.session_state["epi_web_results"] = web_results
        st.session_state["epi_web_errors"] = web_errors
        st.session_state["epi_raw_results"] = raw_results
        st.session_state["epi_web_tables"] = web_tables
        st.session_state["epi_merged_results"] = web_results
        st.session_state["epi_parsed_results"] = web_results
        st.session_state["epi_parse_warnings"] = web_errors.rename(columns={"error": "warning"})

        if web_errors.empty:
            st.success("EPI Web Suite 预测完成。")
        else:
            st.warning(f"预测完成，但有 {len(web_errors)} 个化合物失败。")

    web_results = st.session_state.get("epi_web_results")
    web_errors = st.session_state.get("epi_web_errors")
    web_tables = st.session_state.get("epi_web_tables")
    if web_results is not None:
        st.subheader("网页端预测结果")
        st.dataframe(slim_epi_report_columns(web_results), use_container_width=True)
        render_epi_web_tables(web_tables)
    if web_errors is not None and not web_errors.empty:
        st.subheader("失败记录")
        st.dataframe(web_errors, use_container_width=True)

with tab_fallback:
    st.subheader("EPI Suite 输入文件")
    st.write(
        "如果 EPI Web Suite API 临时不可用，可以下载输入包，手动在网页端计算后，"
        "再回到“解析外部结果”上传结果文件。"
    )
    input_zip = build_input_zip(input_df)
    st.download_button(
        label="下载 EPI Suite 输入包 ZIP",
        data=input_zip,
        file_name="EPISuite_Input_Package.zip",
        mime="application/zip",
    )
    st.write("输入包包含：")
    st.markdown(
        "\n".join(
            [
                "1. `episuite_smiles_only.txt`：每行一个 SMILES，适合复制到 EPI Web Suite。",
                "2. `episuite_named.smi`：SMILES 与化合物名称，用于保留名称映射。",
                "3. `episuite_input.csv`：原始 compound + smiles + 可选 cas 表。",
                "4. `README.txt`：后续上传结果的说明。",
            ]
        )
    )

with tab_parse:
    st.subheader("上传 EPI Suite / EPI Web Suite 结果")
    result_files = st.file_uploader(
        "上传 EPI Suite / EPI Web Suite 结果文件",
        type=["csv", "xlsx", "xls", "txt", "doc"],
        accept_multiple_files=True,
        help="优先推荐 CSV 或 Excel；也支持复制保存的 TXT，以及老版 EPI Suite 的 DOC 文本提取。",
    )

    parsed_frames = []
    warning_frames = []
    if result_files:
        for result_file in result_files:
            try:
                parsed_df, warnings_df = parse_uploaded_result(result_file)
                parsed_frames.append(parsed_df)
                if not warnings_df.empty:
                    warning_frames.append(warnings_df)
            except Exception as exc:
                warning_frames.append(
                    pd.DataFrame(
                        [{"source_file": result_file.name, "warning": f"解析失败：{exc}"}]
                    )
                )

    if parsed_frames:
        parsed_results = pd.concat(parsed_frames, ignore_index=True)
        parse_warnings = (
            pd.concat(warning_frames, ignore_index=True)
            if warning_frames
            else pd.DataFrame(columns=["source_file", "warning"])
        )
        merged_results = merge_results_with_input(input_df, parsed_results)

        st.session_state["epi_parsed_results"] = parsed_results
        st.session_state["epi_merged_results"] = merged_results
        st.session_state["epi_parse_warnings"] = parse_warnings
        st.session_state.pop("epi_raw_results", None)
        st.session_state.pop("epi_web_tables", None)

        st.success("结果文件解析完成。")
        st.subheader("合并后的环境归趋结果")
        st.dataframe(merged_results, use_container_width=True)

        with st.expander("查看原始解析结果", expanded=False):
            st.dataframe(parsed_results, use_container_width=True)

        if not parse_warnings.empty:
            st.warning("部分字段未完全识别，详情见解析警告。")
            st.dataframe(parse_warnings, use_container_width=True)
    else:
        st.info("上传 EPI Suite 结果文件后，会在这里显示结构化解析结果。")

with tab_output:
    st.subheader("下载结果工作簿")
    parsed_results = st.session_state.get("epi_parsed_results")
    merged_results = st.session_state.get("epi_merged_results")
    parse_warnings = st.session_state.get("epi_parse_warnings")
    raw_results = st.session_state.get("epi_raw_results")
    web_tables = st.session_state.get("epi_web_tables")

    workbook_buffer = build_result_workbook(
        input_df,
        parsed_df=parsed_results,
        merged_df=merged_results,
        warnings_df=parse_warnings,
        raw_df=raw_results,
        epi_tables=web_tables,
    )
    workbook_buffer = append_structure_preparation_sheet(workbook_buffer, prepared_input_df)

    st.download_button(
        label="下载 EPI Suite 结果工作簿",
        data=workbook_buffer,
        file_name="EPISuite_Fate_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
